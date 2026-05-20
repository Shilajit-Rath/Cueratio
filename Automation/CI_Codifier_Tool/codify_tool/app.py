#!/usr/bin/env python3
"""
app.py - Flask web app wrapper around codify.py

USAGE
    pip install flask openpyxl
    python app.py
    # then open http://127.0.0.1:5000

The user drops a benchmarking .xlsx file on the page, clicks Convert, and
downloads the codified .xlsx.
"""
from __future__ import annotations

import io
import re
import traceback
from pathlib import Path
from datetime import datetime

from flask import Flask, request, send_file, render_template_string, jsonify, abort
from openpyxl import load_workbook

# Import the codifier module that lives next to this file
from codify import extract_products, build_workbook


app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 64 * 1024 * 1024  # 64 MB upload limit


INDEX_HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>CI Benchmarking → CODIFY Converter</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>
  :root {
    --bg: #f5f7fa;
    --card: #ffffff;
    --primary: #1f4e78;
    --primary-hover: #2e75b6;
    --text: #1f2933;
    --muted: #6b7785;
    --border: #e1e5eb;
    --success: #047857;
    --error: #b91c1c;
    --success-bg: #ecfdf5;
    --error-bg: #fef2f2;
  }
  * { box-sizing: border-box; }
  body {
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
    background: var(--bg);
    color: var(--text);
    margin: 0;
    padding: 40px 20px;
    min-height: 100vh;
  }
  .wrap {
    max-width: 720px;
    margin: 0 auto;
  }
  header h1 {
    margin: 0 0 8px;
    font-size: 28px;
    color: var(--primary);
  }
  header p {
    margin: 0 0 30px;
    color: var(--muted);
    font-size: 15px;
    line-height: 1.5;
  }
  .card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 32px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.04);
  }
  .drop {
    border: 2px dashed #c1c8d2;
    border-radius: 10px;
    padding: 48px 24px;
    text-align: center;
    transition: all 0.15s ease;
    cursor: pointer;
    background: #fafbfc;
  }
  .drop:hover, .drop.over {
    border-color: var(--primary);
    background: #f0f6fb;
  }
  .drop svg {
    width: 48px;
    height: 48px;
    color: var(--primary);
    margin-bottom: 16px;
  }
  .drop-title {
    font-size: 16px;
    font-weight: 600;
    margin-bottom: 6px;
  }
  .drop-sub {
    font-size: 13px;
    color: var(--muted);
  }
  .filename {
    margin-top: 16px;
    padding: 12px 16px;
    background: #f0f6fb;
    border-radius: 8px;
    font-size: 14px;
    display: none;
    word-break: break-all;
  }
  .filename.show { display: block; }
  .filename .remove {
    color: var(--primary);
    cursor: pointer;
    font-weight: 600;
    margin-left: 8px;
  }
  input[type=file] { display: none; }
  .btn {
    margin-top: 20px;
    background: var(--primary);
    color: white;
    border: 0;
    padding: 12px 24px;
    font-size: 15px;
    font-weight: 600;
    border-radius: 8px;
    cursor: pointer;
    width: 100%;
    transition: background 0.15s;
  }
  .btn:hover:not(:disabled) { background: var(--primary-hover); }
  .btn:disabled { opacity: 0.5; cursor: not-allowed; }
  .status {
    margin-top: 20px;
    padding: 14px 16px;
    border-radius: 8px;
    font-size: 14px;
    display: none;
  }
  .status.show { display: block; }
  .status.ok { background: var(--success-bg); color: var(--success); border-left: 4px solid var(--success); }
  .status.err { background: var(--error-bg); color: var(--error); border-left: 4px solid var(--error); }
  .spinner {
    display: inline-block;
    width: 14px;
    height: 14px;
    border: 2px solid #c1c8d2;
    border-top-color: var(--primary);
    border-radius: 50%;
    animation: spin 0.7s linear infinite;
    vertical-align: middle;
    margin-right: 8px;
  }
  @keyframes spin { to { transform: rotate(360deg); } }
  .meta {
    margin-top: 24px;
    padding-top: 20px;
    border-top: 1px solid var(--border);
    font-size: 13px;
    color: var(--muted);
    line-height: 1.6;
  }
  .meta strong { color: var(--text); }
  ul { margin: 6px 0 0; padding-left: 20px; }
  li { margin: 2px 0; }
</style>
</head>
<body>
<div class="wrap">
  <header>
    <h1>CI Benchmarking → CODIFY</h1>
    <p>Drop a benchmarking workbook (Feature Benchmarking + Condition Benchmarking sheets) and download a codified comparison file.</p>
  </header>

  <div class="card">
    <label id="drop" class="drop" for="file">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
        <path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/>
        <polyline points="17 8 12 3 7 8"/>
        <line x1="12" y1="3" x2="12" y2="15"/>
      </svg>
      <div class="drop-title">Drop your .xlsx file here</div>
      <div class="drop-sub">or click to browse</div>
    </label>
    <input type="file" id="file" accept=".xlsx,.xlsm">
    <div id="filename" class="filename"></div>
    <button id="convert" class="btn" disabled>Convert</button>
    <div id="status" class="status"></div>

    <div class="meta">
      <strong>What this does:</strong>
      <ul>
        <li>Auto-detects companies, products and CI conditions from the source.</li>
        <li>Builds the full CODIFY comparison structure (38 sheets).</li>
        <li>Preserves free-text source values in <code>raw_text</code> for traceability.</li>
        <li>Output filename: <code>&lt;your-file&gt;_Codified.xlsx</code>.</li>
      </ul>
    </div>
  </div>
</div>

<script>
const drop = document.getElementById('drop');
const fileInput = document.getElementById('file');
const filename = document.getElementById('filename');
const convertBtn = document.getElementById('convert');
const statusBox = document.getElementById('status');
let selectedFile = null;

function setFile(f) {
  selectedFile = f;
  if (f) {
    filename.innerHTML = '📄 ' + f.name + ' <span class="remove">remove</span>';
    filename.classList.add('show');
    convertBtn.disabled = false;
    statusBox.classList.remove('show');
  } else {
    filename.classList.remove('show');
    convertBtn.disabled = true;
    fileInput.value = '';
  }
}

filename.addEventListener('click', (e) => {
  if (e.target.classList.contains('remove')) {
    setFile(null);
  }
});

fileInput.addEventListener('change', (e) => {
  const f = e.target.files[0];
  if (f) setFile(f);
});

['dragenter', 'dragover'].forEach(ev =>
  drop.addEventListener(ev, (e) => { e.preventDefault(); drop.classList.add('over'); })
);
['dragleave', 'drop'].forEach(ev =>
  drop.addEventListener(ev, (e) => { e.preventDefault(); drop.classList.remove('over'); })
);
drop.addEventListener('drop', (e) => {
  const f = e.dataTransfer.files[0];
  if (f) setFile(f);
});

function showStatus(msg, type) {
  statusBox.innerHTML = msg;
  statusBox.className = 'status show ' + (type || 'ok');
}

convertBtn.addEventListener('click', async () => {
  if (!selectedFile) return;
  convertBtn.disabled = true;
  showStatus('<span class="spinner"></span>Processing...', 'ok');

  const fd = new FormData();
  fd.append('file', selectedFile);

  try {
    const r = await fetch('/convert', { method: 'POST', body: fd });
    if (!r.ok) {
      const err = await r.json().catch(() => ({error: 'Server error ' + r.status}));
      showStatus('❌ ' + (err.error || 'Conversion failed'), 'err');
      convertBtn.disabled = false;
      return;
    }
    const meta = JSON.parse(r.headers.get('X-Codify-Summary') || '{}');
    const blob = await r.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    a.download = selectedFile.name.replace(/\.(xlsx|xlsm)$/i, '') + '_Codified.xlsx';
    document.body.appendChild(a);
    a.click();
    a.remove();
    URL.revokeObjectURL(url);

    let summary = '✓ Conversion complete. Download started.';
    if (meta.companies !== undefined) {
      summary += '<br><br><strong>Summary:</strong>'
        + '<br>Companies: ' + meta.companies
        + ' &nbsp;|&nbsp; Products: ' + meta.products
        + ' &nbsp;|&nbsp; CI conditions: ' + meta.conditions
        + ' &nbsp;|&nbsp; Sheets: ' + meta.sheets;
    }
    showStatus(summary, 'ok');
  } catch (err) {
    showStatus('❌ Network error: ' + err.message, 'err');
  } finally {
    convertBtn.disabled = false;
  }
});
</script>
</body>
</html>
"""


@app.route('/')
def index():
    return render_template_string(INDEX_HTML)


@app.route('/convert', methods=['POST'])
def convert():
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'error': 'No file uploaded.'}), 400
    if not f.filename.lower().endswith(('.xlsx', '.xlsm')):
        return jsonify({'error': 'Please upload an .xlsx or .xlsm file.'}), 400

    try:
        in_bytes = io.BytesIO(f.read())
        wb_in = load_workbook(in_bytes, data_only=True)
    except Exception as e:
        return jsonify({'error': f'Could not read workbook: {e}'}), 400

    try:
        products, cond_data = extract_products(wb_in)
        if not products:
            return jsonify({'error': 'No products were detected. Check the Feature Benchmarking sheet.'}), 400
        wb_out = build_workbook(products, cond_data)
    except Exception as e:
        tb = traceback.format_exc()
        app.logger.error(tb)
        return jsonify({'error': f'Conversion failed: {e}'}), 500

    out_buf = io.BytesIO()
    wb_out.save(out_buf)
    out_buf.seek(0)

    summary = {
        'companies': len({p['company_name'] for p in products}),
        'products': len(products),
        'conditions': len(cond_data['conditions']) if cond_data else 0,
        'sheets': len(wb_out.sheetnames),
    }
    import json
    download_name = re.sub(r'\.(xlsx|xlsm)$', '', f.filename, flags=re.IGNORECASE) + '_Codified.xlsx'

    response = send_file(
        out_buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=download_name,
    )
    response.headers['X-Codify-Summary'] = json.dumps(summary)
    response.headers['Access-Control-Expose-Headers'] = 'X-Codify-Summary'
    return response


@app.route('/health')
def health():
    return jsonify({'status': 'ok', 'timestamp': datetime.utcnow().isoformat()})


if __name__ == '__main__':
    print('━' * 50)
    print('  CI Benchmarking → CODIFY Converter')
    print('  Open: http://127.0.0.1:5000')
    print('  Press Ctrl+C to stop')
    print('━' * 50)
    app.run(host='127.0.0.1', port=5000, debug=False)
