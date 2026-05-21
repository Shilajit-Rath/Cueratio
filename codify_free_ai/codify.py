#!/usr/bin/env python3
"""
codify.py  –  AI-powered CI Benchmarking → CODIFY Converter

Architecture
------------
1. READER      – parses the benchmarking .xlsx into a clean Python structure
2. AI MAPPER   – uses Claude to group features into sheets and extract coded values
3. WRITER      – builds the CODIFY .xlsx (known schemas + auto-generated new sheets)

All AI calls go through the Anthropic /v1/messages endpoint.
ANTHROPIC_API_KEY must be set in the environment (Flask app handles this).
"""
from __future__ import annotations

import json, os, re, time, traceback, urllib.request, urllib.error
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# STYLING
# ─────────────────────────────────────────────────────────────────────────────
ARIAL_10      = Font(name='Arial', size=10)
ARIAL_10_BOLD = Font(name='Arial', size=10, bold=True)
HDR_FILL      = PatternFill('solid', fgColor='BFBFBF')
NO_BORDER     = Border()
HDR_ALIGN     = Alignment(horizontal='left', vertical='top', wrap_text=True)
DATA_ALIGN    = Alignment(horizontal='left', vertical='top', wrap_text=True)
HEADER_ROW    = 5

# ─────────────────────────────────────────────────────────────────────────────
# KNOWN SCHEMAS  (mirrors CODIFY_COMPARISON v2.9)
# ─────────────────────────────────────────────────────────────────────────────
ID_COLS = ['Variant ID', 'Variant Name', 'Product Name', 'Company Name']

KNOWN_SCHEMAS: dict[str, list[str]] = {
    '01_Company_Master': ['Company ID', 'Company Name'],
    '02_Product_Master': ['Product ID', 'Product Name'],
    '03_Variant_Master': ID_COLS + ['Plan Type'],
    '04_Entry_Age': ID_COLS + [
        'min_entry_age_value','min_entry_age_unit','min_entry_age_in_days',
        'min_entry_age_in_months','min_entry_age_in_years',
        'min_child_age_value','min_child_age_unit','min_child_age_in_days',
        'min_child_age_in_months','min_child_age_in_years',
        'max_entry_age_value','max_entry_age_unit','max_entry_age_in_days',
        'max_entry_age_in_months','max_entry_age_in_years',
        'max_child_age_value','max_child_age_unit','max_child_age_in_days',
        'max_child_age_in_months','max_child_age_in_years',
        'max_renewal_age_value','max_renewal_age_unit','max_renewal_age_in_days',
        'max_renewal_age_in_months','max_renewal_age_in_years',
        'special_conditions','raw_text',
    ],
    '05_Policy_Term': ID_COLS + [
        'policy_term_value','policy_term_unit','policy_term_in_days',
        'policy_term_in_months','policy_term_in_years',
        'special_conditions','raw_text',
    ],
    '06_Coverage': ID_COLS + [
        'coverage_max_age_value','coverage_max_age_unit','coverage_max_age_in_days',
        'coverage_max_age_in_months','coverage_max_age_in_years',
        'renewal_max_age_value','renewal_max_age_unit','renewal_max_age_in_days',
        'renewal_max_age_in_months','renewal_max_age_in_years',
        'coverage_linked_to_base_plan','requires_continuous_renewal',
        'coverage_termination_condition','lifetime_renewal_eligibility',
        'lifetime_renewal_condition','special_conditions','raw_text',
    ],
    '07A_Waiting_Period_Standard': ID_COLS + [
        'waiting_period_type','waiting_period_value','waiting_period_unit',
        'waiting_period_in_days','waiting_period_in_months','waiting_period_in_years',
        'waiting_period_conditions','special_conditions','raw_text',
    ],
    '07B_Waiting_Period_Accident': ID_COLS + [
        'waiting_period_type','waiting_period_value','waiting_period_unit',
        'waiting_period_in_days','waiting_period_in_months','waiting_period_in_years',
        'waiting_period_conditions','special_conditions','raw_text',
    ],
    '07C_Waiting_Period_Specific': ID_COLS + [
        'waiting_period_type','waiting_period_value','waiting_period_unit',
        'waiting_period_in_days','waiting_period_in_months','waiting_period_in_years',
        'waiting_period_conditions','special_conditions','raw_text',
    ],
    '07D_Waiting_Period_Special(90D)': ID_COLS + [
        'waiting_period_type','waiting_period_value','waiting_period_unit',
        'waiting_period_in_days','waiting_period_in_months','waiting_period_in_years',
        'waiting_period_conditions','special_conditions','raw_text',
    ],
    '07E_Waiting_Period_Other(12MWP)': ID_COLS + [
        'waiting_period_type','waiting_period_value','waiting_period_unit',
        'waiting_period_in_days','waiting_period_in_months','waiting_period_in_years',
        'waiting_period_conditions','special_conditions','raw_text',
    ],
    '08_Policy_Year_Benefit_Amt': ID_COLS + [
        'policy_year_max_benefit','policy_year_max_benefit_currency',
        'si_applicability','special_conditions','raw_text',
    ],
    '09_Hospital_Stay_Benefit_Amt': ID_COLS + [
        'per_hospital_stay_max_benefit','per_hospital_stay_max_benefit_currency',
        'si_applicability','special_conditions','raw_text',
    ],
    '10_Area_Covered': ID_COLS + [
        'coverage_area_type','exclusion_coverage_area','special_conditions','raw_text',
    ],
    '11_Payment_Frequency': ID_COLS + [
        'annual_mode','semi_annual_mode','quarterly_mode','monthly_mode',
        'mode_of_payment','payment_requires_base_plan','special_conditions','raw_text',
    ],
    '12_Family_Discounts': ID_COLS + [
        'has_family_discount','family_discount_type','family_discount_percentage',
        'discount_basis','special_conditions','raw_text',
    ],
    '13_Room_Board_Inpatient': ID_COLS + [
        'benefit_name','room_type_applicablity','room_type','room_rent_amount',
        'room_rent_amount_unit','room_rent_max_days','room_rent_payable_mode',
        'si_applicability','special_conditions','raw_text',
    ],
    '14_ICU_Room_Inpatient': ID_COLS + [
        'benefit_name','room_type_applicablity','room_type','icu_room_rent_amount',
        'icu_room_rent_amount_unit','icu_room_rent_max_days','icu_rent_type',
        'icu_room_rent_payable_mode','si_applicability','special_conditions','raw_text',
    ],
    '15_Medical_Services_Fees': ID_COLS + [
        'benefit_name','benefit_level','sub_benefit_name','medical_services_payable_mode',
        'medical_services_amount_unit','medical_services_amount_limit',
        'si_applicability','special_conditions','raw_text',
    ],
    '16_IPD_Physician_Fees': ID_COLS + [
        'benefit_name','physician_fees_applicability','physician_fees_amount',
        'physician_fees_amount_unit','physician_fees_payable_mode',
        'si_applicability','special_conditions','raw_text',
    ],
    '17_Medical_Expenses_Surgery_Pro': ID_COLS + [
        'benefit_name','benefit_level','sub_benefit_name','amount_limit','amount_unit',
        'payable_mode','si_applicability','special_conditions','raw_text',
    ],
    '18_Day_Surgery_Major': ID_COLS + [
        'benefit_name','day_surgery_applicability','day_surgery_amount',
        'day_surgery_amount_unit','day_surgery_payable_mode',
        'si_applicability','special_conditions','raw_text',
    ],
    '19_Diagnostic_Services_IPD': ID_COLS + [
        'benefit_name','benefit_level','sub_benefit_name','amount_limit','amount_unit',
        'payable_mode','si_applicability','special_conditions','raw_text',
    ],
    '20_OPD_Accident_24_Hours': ID_COLS + [
        'benefit_name','opd_accident_applicability','opd_accident_amount_value',
        'opd_accident_amount_unit','opd_accident_payable_mode',
        'si_applicability','special_conditions','raw_text',
    ],
    '21_Rehabilitation_Expenses_IPD': ID_COLS + [
        'benefit_name','rehab_applicability','rehab_amount_value','rehab_amount_unit',
        'rehab_payable_mode','si_applicability','special_conditions','raw_text',
    ],
    '22_Dialysis_Treatment': ID_COLS + [
        'benefit_name','dialysis_applicability','dialysis_type',
        'dialysis_amount_value','dialysis_amount_unit','dialysis_payable_mode',
        'si_applicability','special_conditions','raw_text',
    ],
    '23_Cancer_Treatment_Radiation': ID_COLS + [
        'benefit_name','cancer_radiation_applicability','treatment_type',
        'cancer_radiation_amount_value','cancer_radiation_amount_unit',
        'cancer_radiation_payable_mode','si_applicability','special_conditions','raw_text',
    ],
    '24_Chemo_Treatment_Expenses': ID_COLS + [
        'benefit_name','chemo_applicability','chemotherapy_type',
        'chemo_amount_value','chemo_amount_unit','chemo_payable_mode',
        'si_applicability','special_conditions','raw_text',
    ],
    '25_Emergency_Ambulance_Services': ID_COLS + [
        'benefit_name','ambulance_applicability','ambulance_amount_value',
        'ambulance_amount_unit','ambulance_payable_mode',
        'si_applicability','special_conditions','raw_text',
    ],
    '26_Minor_Surgery_Expenses': ID_COLS + [
        'benefit_name','minor_surgery_applicability','minor_surgery_amount_value',
        'minor_surgery_amount_unit','minor_surgery_payable_mode',
        'si_applicability','special_conditions','raw_text',
    ],
    '27_OPD_Medical_Services': ID_COLS + [
        'benefit_name','opd_applicability','opd_treatment_type',
        'opd_amount_value','opd_amount_unit','opd_payable_mode',
        'si_applicability','special_conditions','raw_text',
    ],
    '28_Critical_Illness_Benefits': ID_COLS + [
        'benefit_name','ci_benefit_type','ci_trigger_condition',
        'ci_amount_value','ci_amount_type','ci_payable_mode',
        'ci_multiple_claims_allowed','ci_payout_structure',
        'si_applicability','special_conditions','raw_text',
    ],
    '29_Misc_&_Extra_Benefits': ID_COLS + [
        'benefit_name','sub_benefit_name','applicability','payable_mode',
        'amount_value','amount_unit','max_limit_value','max_limit_unit',
        'si_applicability','special_conditions','raw_text',
    ],
    '30_Deductible': ID_COLS + [
        'benefit_name','deductible_applicability','deductible_type',
        'deductible_amount_value','deductible_amount_unit',
        'deductible_frequency','special_conditions','raw_text',
    ],
    '31_Copayment': ID_COLS + [
        'benefit_name','copayment_applicability','copayment_type',
        'copayment_value','copayment_unit','copayment_trigger',
        'special_conditions','raw_text',
    ],
    '32_No_Claim_Bonus': ID_COLS + [
        'benefit_name','ncb_type','ncb_rate_value','ncb_rate_unit',
        'ncb_max_limit','ncb_max_limit_unit','ncb_reset_condition',
        'special_conditions','raw_text',
    ],
}


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — READER
# ─────────────────────────────────────────────────────────────────────────────

def read_benchmarking(wb) -> dict:
    """Parse benchmarking workbook into clean Python structure."""
    SKIP = {'features','feature','benefits','benefit','s.no.','sno','category',
            'no.','plans','plan','conditions','condition','s.no',
            'individual_product_detail','product_brochure','unit','unit - baht'}

    # Find the Feature Benchmarking sheet
    ws = None
    for sn in wb.sheetnames:
        candidate = wb[sn]
        for r in range(1, 6):
            for c in range(1, 6):
                v = candidate.cell(row=r, column=c).value
                if isinstance(v, str) and v.strip().lower() in {'company','companies'}:
                    ws = candidate
                    break
            if ws: break
        if ws: break

    if ws is None:
        raise RuntimeError("Could not find a Feature Benchmarking sheet.")

    # Locate structural rows
    company_row = product_row = label_col = feat_start = None
    for r in range(1, 10):
        for c in range(1, 6):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str): continue
            vs = v.strip().lower()
            if vs in {'company','companies','insurance company'} and company_row is None:
                company_row, label_col = r, c
            elif vs in {'products','product','product name'} and product_row is None:
                product_row = r

    if not company_row or not product_row:
        raise RuntimeError("Could not locate Company/Products rows.")

    start_c = (label_col + 1) if label_col else 2
    first_dc = last_dc = None
    for c in range(start_c, ws.max_column + 1):
        v = ws.cell(row=company_row, column=c).value
        if v and str(v).strip() not in {'', 'NA', 'N/A'}:
            if first_dc is None: first_dc = c
            last_dc = c

    if not first_dc:
        raise RuntimeError("No product columns found.")

    # Find feature start row
    for r in range(product_row + 1, min(product_row + 8, ws.max_row + 1)):
        for c in range(1, 4):
            v = ws.cell(row=r, column=c).value
            if (isinstance(v, str) and v.strip()
                    and v.strip().lower() not in SKIP
                    and not v.strip().isdigit()
                    and not v.lower().startswith('http')):
                feat_start = r
                break
        if feat_start: break

    if not feat_start:
        feat_start = product_row + 2

    # Detect feature-name and category columns
    counts = {1: 0, 2: 0, 3: 0}
    for r in range(feat_start, min(feat_start + 40, ws.max_row + 1)):
        for c in (1, 2, 3):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() and not v.strip().isdigit():
                counts[c] += 1
    sorted_c = sorted(counts.items(), key=lambda x: -x[1])
    top = [c for c, n in sorted_c if n > 3]
    if len(top) >= 2:
        cat_col, feat_col = sorted(top[:2])
    elif len(top) == 1:
        feat_col, cat_col = top[0], None
    else:
        feat_col, cat_col = 3, 2

    # Build products
    products = []
    for c in range(first_dc, last_dc + 1):
        company = ws.cell(row=company_row, column=c).value
        product = ws.cell(row=product_row, column=c).value
        if not company or not product: continue
        cs = str(company).strip()
        ps = str(product).strip().rstrip('\t').strip()
        if not cs or not ps or cs.lower() in {'na','n/a'} or ps.lower() in {'na','n/a'}:
            continue
        products.append({'company': cs, 'product': ps, 'col_index': c, 'features': {}})

    if not products:
        raise RuntimeError("No products detected.")

    # Collect feature values
    feature_order, categories = [], {}
    for r in range(feat_start, ws.max_row + 1):
        fn_raw = ws.cell(row=r, column=feat_col).value
        if not isinstance(fn_raw, str) or not fn_raw.strip(): continue
        fn = fn_raw.strip()
        if fn.lower() in SKIP: continue
        if fn not in categories:
            feature_order.append(fn)
            cat = ws.cell(row=r, column=cat_col).value if cat_col else None
            categories[fn] = str(cat).strip() if cat and str(cat).strip() else 'General'
        for p in products:
            v = ws.cell(row=r, column=p['col_index']).value
            p['features'][fn] = str(v).strip() if v is not None else 'NA'

    return {'products': products, 'feature_order': feature_order, 'categories': categories}


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — FREE / LOCAL AI MAPPER
# ─────────────────────────────────────────────────────────────────────────────

# Default free AI backend: Ollama running locally on your laptop.
# Install Ollama, then run: ollama pull llama3.1:8b
OLLAMA_URL   = os.environ.get('OLLAMA_URL', 'http://127.0.0.1:11434/api/chat')
OLLAMA_MODEL = os.environ.get('OLLAMA_MODEL', 'llama3.1:8b')

DEFAULT_NEW_SCHEMA = [
    'benefit_name', 'applicability', 'amount_value', 'amount_unit',
    'limit_value', 'limit_unit', 'special_conditions', 'raw_text'
]


def _call_ai(system: str, user: str, max_tokens: int = 4000) -> str:
    """Call a free local Ollama model and return text."""
    payload = json.dumps({
        'model': OLLAMA_MODEL,
        'stream': False,
        'options': {
            'temperature': 0.1,
            'num_predict': max_tokens,
        },
        'messages': [
            {'role': 'system', 'content': system},
            {'role': 'user', 'content': user},
        ],
    }).encode('utf-8')

    req = urllib.request.Request(
        OLLAMA_URL,
        data=payload,
        headers={'Content-Type': 'application/json'},
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=240) as resp:
            data = json.loads(resp.read().decode('utf-8'))
            return data.get('message', {}).get('content', '')
    except urllib.error.URLError as e:
        raise RuntimeError(
            'Could not connect to Ollama. Start it with `ollama serve`, '
            'then run `ollama pull llama3.1:8b`. Original error: ' + str(e)
        ) from e


def _parse_json(text: str):
    """Parse model output even if it accidentally includes markdown or extra text."""
    text = text.strip()
    text = re.sub(r'^```(?:json)?\s*', '', text, flags=re.MULTILINE)
    text = re.sub(r'\s*```$', '', text, flags=re.MULTILINE)
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        # Last-resort extraction of first JSON object.
        m = re.search(r'\{.*\}', text, flags=re.DOTALL)
        if not m:
            raise
        return json.loads(m.group(0))


def _slug_col(name: str) -> str:
    name = str(name or '').strip().lower()
    name = re.sub(r'[^a-z0-9]+', '_', name).strip('_')
    return name or 'value'


def _safe_sheet_name(name: str, used: set | None = None) -> str:
    """Excel-safe sheet name, max 31 chars, with stable readable format."""
    used = used or set()
    name = str(name or '33_Extra_Feature').strip()
    name = re.sub(r'[\\/*?:\[\]]+', '_', name)
    name = re.sub(r'\s+', '_', name)
    name = name[:31].strip('_') or '33_Extra_Feature'
    base = name
    i = 2
    while name in used:
        suffix = f'_{i}'
        name = base[:31-len(suffix)] + suffix
        i += 1
    return name


def _renumber_new_sheets(feature_to_sheet: dict, new_schemas: dict) -> tuple[dict, dict]:
    """Ensure new sheets start at 33 and are Excel-safe."""
    known = set(KNOWN_SCHEMAS)
    mapping = {}
    used = set(known)
    next_no = 33

    for old in sorted({v for v in feature_to_sheet.values() if v and v not in known} | set(new_schemas.keys())):
        # Keep readable name after any leading number.
        label = re.sub(r'^\d+[_\s-]*', '', str(old)).strip() or 'Extra_Feature'
        candidate = _safe_sheet_name(f'{next_no:02d}_{label}', used)
        mapping[old] = candidate
        used.add(candidate)
        next_no += 1

    cleaned_feature_to_sheet = {
        fn: mapping.get(sn, sn) for fn, sn in feature_to_sheet.items() if sn
    }
    cleaned_new_schemas = {
        mapping.get(sn, sn): cols for sn, cols in new_schemas.items()
    }
    return cleaned_feature_to_sheet, cleaned_new_schemas


def validate_new_schemas(new_schemas: dict) -> dict:
    """Python guardrail: AI may suggest columns, but this controls final structure."""
    cleaned = {}
    identity = {'variant_id', 'variant_name', 'product_name', 'company_name'}

    for sheet_name, cols in (new_schemas or {}).items():
        if not isinstance(cols, list):
            cols = DEFAULT_NEW_SCHEMA[:]

        normalized = []
        for c in cols:
            c = _slug_col(c)
            if c and c not in identity and c not in normalized:
                normalized.append(c)

        business_cols = [
            c for c in normalized
            if c not in {'special_conditions', 'raw_text'}
        ][:8]
        if not business_cols:
            business_cols = DEFAULT_NEW_SCHEMA[:-2]

        cleaned[sheet_name] = business_cols + ['special_conditions', 'raw_text']

    return cleaned


def _fallback_plan(bench: dict) -> dict:
    """Rule-based fallback if local AI fails to produce usable JSON."""
    feature_to_sheet = {}
    new_schemas = {}
    next_no = 33

    rules = [
        (r'entry.*age|min.*age|max.*age|renewal.*age', '04_Entry_Age'),
        (r'policy.*term|term', '05_Policy_Term'),
        (r'coverage|lifetime|renewal', '06_Coverage'),
        (r'waiting.*accident|accident.*waiting', '07B_Waiting_Period_Accident'),
        (r'waiting.*specific|specific.*waiting', '07C_Waiting_Period_Specific'),
        (r'waiting.*90|90.*waiting', '07D_Waiting_Period_Special(90D)'),
        (r'waiting|pre.?existing|moratorium', '07A_Waiting_Period_Standard'),
        (r'payment|frequency|mode', '11_Payment_Frequency'),
        (r'family.*discount|discount', '12_Family_Discounts'),
        (r'room|board|inpatient', '13_Room_Board_Inpatient'),
        (r'icu', '14_ICU_Room_Inpatient'),
        (r'physician|doctor', '16_IPD_Physician_Fees'),
        (r'surgery|surgical', '17_Medical_Expenses_Surgery_Pro'),
        (r'ambulance', '25_Emergency_Ambulance_Services'),
        (r'critical|ci_', '28_Critical_Illness_Benefits'),
        (r'deductible', '30_Deductible'),
        (r'copay|co.?pay', '31_Copayment'),
        (r'no.?claim|bonus|ncb', '32_No_Claim_Bonus'),
    ]

    for fn in bench['feature_order']:
        low = fn.lower()
        mapped = None
        for pat, sheet in rules:
            if re.search(pat, low):
                mapped = sheet
                break
        if not mapped:
            label = re.sub(r'[^A-Za-z0-9]+', '_', fn).strip('_')[:24] or 'Extra_Feature'
            mapped = f'{next_no:02d}_{label}'
            next_no += 1
            new_schemas[mapped] = DEFAULT_NEW_SCHEMA[:]
        feature_to_sheet[fn] = mapped

    return {'feature_to_sheet': feature_to_sheet, 'new_sheet_schemas': new_schemas}


def ai_plan_sheets(bench: dict) -> dict:
    """Ask free local AI to map each feature → sheet and define schemas for new sheets."""
    known_list = '\n'.join(
        f'  - {k}' for k in KNOWN_SCHEMAS
        if k not in ('01_Company_Master', '02_Product_Master', '03_Variant_Master')
    )
    feature_summary = '\n'.join(
        f'  [{bench["categories"].get(f, "General")}] {f}'
        for f in bench['feature_order']
    )
    system = (
        'You are a structured data architect for insurance product comparison. '
        'Map benchmarking features to controlled CODIFY sheet names. '
        'Return only valid JSON. Do not use markdown.'
    )
    user = f"""Map each feature below to the correct CODIFY sheet.

KNOWN SHEETS (prefer these when a feature clearly fits):
{known_list}

STRICT RULES:
- Use known sheets whenever possible.
- Product_Category, Target_Market, Product_Positioning -> "03_Variant_Master".
- Waiting_Period features -> choose the most appropriate 07A-07E sheet.
- Unknown features -> assign a NEW sheet name like "33_BenefitName".
- Related feature groups belong to one sheet; unrelated features get separate sheets.
- For new sheet columns: snake_case, 4-8 business columns max.
- Every new schema must end with "special_conditions", "raw_text".
- Return JSON only.

FEATURES:
{feature_summary}

Return this exact JSON shape:
{{
  "feature_to_sheet": {{"feature_name": "sheet_name"}},
  "new_sheet_schemas": {{"new_sheet_name": ["col1", "col2", "special_conditions", "raw_text"]}},
  "feature_reasons": {{"feature_name": "short reason"}}
}}
"""
    try:
        raw = _call_ai(system, user, max_tokens=3500)
        plan = _parse_json(raw)
        if not isinstance(plan, dict) or not plan.get('feature_to_sheet'):
            raise ValueError('AI returned empty plan')
        return plan
    except Exception:
        traceback.print_exc()
        return _fallback_plan(bench)


def ai_extract_sheet(sheet_name: str, columns: list, feature_names: list,
                     products: list) -> list:
    """Ask free local AI to fill one sheet across all products."""
    feat_table = {
        fn: {p['product']: p['features'].get(fn, 'NA') for p in products}
        for fn in feature_names
    }
    product_ids = [{'product': p['product'], 'company': p['company']} for p in products]
    cols_desc = ', '.join(columns[4:])   # skip ID cols

    system = (
        'You are an expert insurance data analyst. Extract and codify insurance product '
        'feature values into a structured comparison table. Return only valid JSON.'
    )
    user = f"""Fill the CODIFY sheet "{sheet_name}".

PRODUCTS:
{json.dumps(product_ids, indent=2)}

RAW FEATURE VALUES:
{json.dumps(feat_table, indent=2)}

COLUMNS TO FILL after identity columns:
{cols_desc}

EXTRACTION RULES:
- One JSON object per product in "rows" array unless this is a multi-benefit sheet.
- "product" key must match the product name exactly.
- "raw_text": paste the most relevant raw source text verbatim.
- "special_conditions": important caveat not captured elsewhere, else null.
- Use "NA" when a product genuinely does not have that benefit.
- Amount columns: numeric value only; put unit in the corresponding _unit column.
- Yes/No columns: "Yes" or "No" only.
- Keep values concise and factual.

Return only:
{{"rows": [{{"product": "name", "col": "value"}}]}}
"""
    raw = _call_ai(system, user, max_tokens=4500)
    parsed = _parse_json(raw)
    rows = parsed.get('rows', []) if isinstance(parsed, dict) else []
    return rows if isinstance(rows, list) else []


def _validate_rows(rows: list, columns: list, products: list, sheet_name: str) -> list:
    """Keep only controlled columns and make sure raw_text exists."""
    allowed = set(columns[4:]) | {'product'}
    cleaned = []
    product_names = {p['product'] for p in products}

    for row in rows or []:
        if not isinstance(row, dict):
            continue
        out = {k: row.get(k) for k in allowed if k in row}
        pname = out.get('product')
        if pname not in product_names:
            # Try a forgiving match.
            for p in product_names:
                if str(p).strip().lower() == str(pname).strip().lower():
                    out['product'] = p
                    break
        if not out.get('raw_text'):
            out['raw_text'] = 'NA'
        cleaned.append(out)

    return cleaned


def run_ai_pipeline(bench: dict, progress_cb=None) -> dict:
    """Full AI pipeline: bench dict → sheet_rows dict."""
    products = bench['products']

    if progress_cb: progress_cb(f'Planning sheet structure with free local AI ({OLLAMA_MODEL})…')
    plan = ai_plan_sheets(bench)
    feature_to_sheet: dict = plan.get('feature_to_sheet', {}) or {}
    new_schemas: dict = plan.get('new_sheet_schemas', {}) or {}

    feature_to_sheet, new_schemas = _renumber_new_sheets(feature_to_sheet, new_schemas)
    new_schemas = validate_new_schemas(new_schemas)

    # Group features by target sheet
    sheet_features: dict[str, list] = {}
    for fn in bench['feature_order']:
        sn = feature_to_sheet.get(fn)
        if sn:
            sheet_features.setdefault(sn, []).append(fn)

    master = {'01_Company_Master', '02_Product_Master', '03_Variant_Master'}
    sheets_to_fill = {sn for sn in sheet_features if sn not in master and sheet_features[sn]}

    sheet_rows: dict[str, list] = {}
    for sn in sorted(sheets_to_fill):
        if progress_cb: progress_cb(f'Extracting {sn}…')
        cols = KNOWN_SCHEMAS.get(sn) or (ID_COLS + new_schemas.get(sn, DEFAULT_NEW_SCHEMA))
        try:
            rows = ai_extract_sheet(sn, cols, sheet_features[sn], products)
            rows = _validate_rows(rows, cols, products, sn)
        except Exception:
            traceback.print_exc()
            rows = []
        sheet_rows[sn] = rows
        time.sleep(0.15)

    audit_rows = []
    reasons = plan.get('feature_reasons', {}) or {}
    for fn in bench['feature_order']:
        sn = feature_to_sheet.get(fn, '')
        samples = []
        for p in products[:3]:
            v = p['features'].get(fn, 'NA')
            if v and v != 'NA':
                samples.append(f"{p['product']}: {v}")
        audit_rows.append({
            'feature_name': fn,
            'category': bench['categories'].get(fn, 'General'),
            'mapped_sheet': sn,
            'known_or_new': 'Known' if sn in KNOWN_SCHEMAS else 'New',
            'reason': reasons.get(fn, ''),
            'raw_sample': ' | '.join(samples)[:1000],
        })

    return {'sheet_plan': plan, 'sheet_rows': sheet_rows,
            'products': [{'company': p['company'], 'product': p['product']} for p in products],
            'new_schemas': new_schemas,
            'audit_rows': audit_rows}


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — WRITER
# ─────────────────────────────────────────────────────────────────────────────

def _init_sheet(wb: Workbook, name: str, headers: list):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=HEADER_ROW, column=c_idx, value=h)
        cell.font = ARIAL_10_BOLD
        cell.fill = HDR_FILL
        cell.alignment = HDR_ALIGN
        cell.border = NO_BORDER
    return ws


def _write_row(ws, row_num: int, values: list, n_cols: int):
    for c_idx in range(1, n_cols + 1):
        v = values[c_idx - 1] if c_idx <= len(values) else None
        cell = ws.cell(row=row_num, column=c_idx, value=v)
        cell.font = ARIAL_10
        cell.alignment = DATA_ALIGN
        cell.border = NO_BORDER


def _short_name(full: str) -> str:
    stop = {'life','insurance','corporation','company','inc','ltd','incorporated',
            'assurance','and','of','the','philippines','public','limited','pvt',
            'co','general'}
    words = [w for w in re.split(r'[\s,\.]+', full) if w and w.lower() not in stop]
    return ' '.join(words[:4]) or full[:30]


def build_workbook(pipeline_result: dict) -> Workbook:
    products   = pipeline_result['products']
    sheet_rows = pipeline_result['sheet_rows']
    new_schemas = pipeline_result.get('new_schemas', {})
    audit_rows = pipeline_result.get('audit_rows', [])

    wb = Workbook()
    wb.remove(wb.active)

    # Assign company IDs
    company_id: dict[str, str] = {}
    ci = 1
    for p in products:
        if p['company'] not in company_id:
            company_id[p['company']] = f'C{ci:02d}'
            ci += 1

    # Assign product IDs
    product_id: dict[tuple, str] = {}
    prod_count: dict[str, int] = {}
    for p in products:
        key = (p['company'], p['product'])
        if key not in product_id:
            cid = company_id[p['company']]
            prod_count[cid] = prod_count.get(cid, 0) + 1
            product_id[key] = f"{cid}-P{prod_count[cid]:02d}"

    # Enrich products
    for p in products:
        key = (p['company'], p['product'])
        p['variant_id']    = f"{product_id[key]}-V01"
        p['variant_name']  = p['product']
        p['company_short'] = _short_name(p['company'])
        p['product_id']    = product_id[key]
        p['company_id']    = company_id[p['company']]

    # 00 AI Mapping Audit — explains every AI grouping decision
    audit_headers = ['Feature Name', 'Category', 'Mapped Sheet', 'Known/New', 'Reason', 'Raw Sample']
    ws = _init_sheet(wb, '00_AI_Mapping_Audit', audit_headers)
    r = HEADER_ROW + 1
    for ar in audit_rows:
        _write_row(ws, r, [
            ar.get('feature_name'), ar.get('category'), ar.get('mapped_sheet'),
            ar.get('known_or_new'), ar.get('reason'), ar.get('raw_sample')
        ], len(audit_headers))
        r += 1

    # 01 Company Master
    ws = _init_sheet(wb, '01_Company_Master', KNOWN_SCHEMAS['01_Company_Master'])
    seen_co = set()
    r = HEADER_ROW + 1
    for p in products:
        if p['company'] not in seen_co:
            _write_row(ws, r, [p['company_id'], p['company']], 2)
            seen_co.add(p['company']); r += 1

    # 02 Product Master
    ws = _init_sheet(wb, '02_Product_Master', KNOWN_SCHEMAS['02_Product_Master'])
    seen_pr = set()
    r = HEADER_ROW + 1
    for p in products:
        key = (p['company'], p['product'])
        if key not in seen_pr:
            _write_row(ws, r, [p['product_id'], p['product']], 2)
            seen_pr.add(key); r += 1

    # 03 Variant Master
    ws = _init_sheet(wb, '03_Variant_Master', KNOWN_SCHEMAS['03_Variant_Master'])
    for i, p in enumerate(products, 1):
        _write_row(ws, HEADER_ROW + i,
                   [p['variant_id'], p['variant_name'], p['product'], p['company'], 'Standalone'],
                   len(KNOWN_SCHEMAS['03_Variant_Master']))

    # Lookup: product name → enriched product dict
    prod_lookup = {}
    for p in products:
        prod_lookup[p['product']] = p
        prod_lookup[p['product'].rstrip('\t').strip()] = p

    def id_cols(product_name: str) -> list:
        p = prod_lookup.get(product_name) or prod_lookup.get(product_name.strip())
        if p:
            return [p['variant_id'], p['variant_name'], p['product'], p['company']]
        return [None, product_name, product_name, None]

    # Write all AI-filled sheets
    known_order = [k for k in KNOWN_SCHEMAS
                   if k not in ('01_Company_Master','02_Product_Master','03_Variant_Master')]
    new_order   = sorted(new_schemas.keys())
    all_to_write = [s for s in known_order if s in sheet_rows] + \
                   [s for s in new_order   if s in sheet_rows]

    for sn in all_to_write:
        cols = KNOWN_SCHEMAS.get(sn) or (ID_COLS + new_schemas.get(sn, ['raw_text']))
        rows = sheet_rows.get(sn, [])
        ws = _init_sheet(wb, sn, cols)
        n_cols = len(cols)
        data_row = HEADER_ROW + 1
        for row_obj in rows:
            if not isinstance(row_obj, dict): continue
            pname = row_obj.get('product', '')
            vals = id_cols(pname) + [row_obj.get(c) for c in cols[4:]]
            _write_row(ws, data_row, vals, n_cols)
            data_row += 1

    # Auto-fit columns
    for ws in wb.worksheets:
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 60))
            ws.column_dimensions[col_letter].width = max(12, max_len + 2)

    return wb


# ─────────────────────────────────────────────────────────────────────────────
# PUBLIC ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def convert(wb_in, progress_cb=None) -> tuple:
    """
    Main entry: benchmarking wb → (codify wb, summary dict).
    progress_cb(message) called at each stage if provided.
    """
    if progress_cb: progress_cb('Reading benchmarking sheet…')
    bench = read_benchmarking(wb_in)

    result = run_ai_pipeline(bench, progress_cb=progress_cb)

    if progress_cb: progress_cb('Building output workbook…')
    wb_out = build_workbook(result)

    summary = {
        'companies':  len({p['company'] for p in bench['products']}),
        'products':   len({p['product'] for p in bench['products']}),
        'features':   len(bench['feature_order']),
        'sheets':     len(wb_out.sheetnames),
        'new_sheets': list(result.get('new_schemas', {}).keys()),
    }
    return wb_out, summary
