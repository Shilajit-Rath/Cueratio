#!/usr/bin/env python3
"""
app.py - Flask web app wrapper around codify.py

USAGE
    pip install flask openpyxl
    python app.py
    # then open http://127.0.0.1:5000

The user drops a benchmarking .xlsx file on the page, clicks Convert, and
downloads the codified .xlsx.
"""
from __future__ import annotations

import io
import re
import traceback
from pathlib import Path
from datetime import datetime

from flask import Flask, request, send_file, render_template_string, jsonify, abort
from openpyxl import load_workbook

# Import the codifier module that lives next to this file
from codify import extract_products, build_workbook


app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 64 * 1024 * 1024  # 64 MB upload limit


INDEX_HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>CI Benchmarking → CODIFY Converter</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

  :root {
    --bg: #f0f2f5;
    --surface: #ffffff;
    --surface-2: #f7f8fa;
    --primary: #1a4f8a;
    --primary-light: #e8f1fb;
    --primary-hover: #1560a8;
    --text: #111827;
    --text-muted: #6b7280;
    --text-soft: #9ca3af;
    --border: #e5e7eb;
    --border-strong: #d1d5db;
    --success: #065f46;
    --success-bg: #ecfdf5;
    --success-border: #6ee7b7;
    --error: #991b1b;
    --error-bg: #fff1f2;
    --error-border: #fca5a5;
    --warning: #92400e;
    --warning-bg: #fffbeb;
    --radius: 10px;
    --radius-sm: 6px;
    --shadow: 0 1px 4px rgba(0,0,0,0.06), 0 4px 16px rgba(0,0,0,0.04);
  }

  body {
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    background: var(--bg);
    color: var(--text);
    min-height: 100vh;
    padding: 48px 20px 64px;
    line-height: 1.5;
    -webkit-font-smoothing: antialiased;
  }

  .page {
    max-width: 640px;
    margin: 0 auto;
  }

  /* ── Header ── */
  .header {
    display: flex;
    align-items: center;
    gap: 14px;
    margin-bottom: 32px;
  }
  .header-icon {
    width: 44px;
    height: 44px;
    background: var(--primary);
    border-radius: var(--radius-sm);
    display: flex;
    align-items: center;
    justify-content: center;
    flex-shrink: 0;
  }
  .header-icon svg { width: 22px; height: 22px; stroke: #fff; }
  .header-text h1 {
    font-size: 20px;
    font-weight: 700;
    color: var(--text);
    letter-spacing: -0.3px;
  }
  .header-text p {
    font-size: 13.5px;
    color: var(--text-muted);
    margin-top: 2px;
  }

  /* ── Card ── */
  .card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    box-shadow: var(--shadow);
    overflow: hidden;
  }
  .card-body { padding: 28px; }

  /* ── Drop zone ── */
  .dropzone {
    border: 2px dashed var(--border-strong);
    border-radius: var(--radius);
    padding: 40px 24px;
    text-align: center;
    cursor: pointer;
    transition: border-color 0.15s, background 0.15s;
    background: var(--surface-2);
    position: relative;
  }
  .dropzone:hover, .dropzone.over {
    border-color: var(--primary);
    background: var(--primary-light);
  }
  .dropzone.has-file {
    border-style: solid;
    border-color: var(--primary);
    background: var(--primary-light);
    padding: 20px 24px;
  }
  .dropzone input[type=file] {
    position: absolute; inset: 0; opacity: 0; cursor: pointer; width: 100%; height: 100%;
  }
  .dz-idle { display: flex; flex-direction: column; align-items: center; gap: 10px; }
  .dz-icon {
    width: 48px; height: 48px;
    background: #dbeafe;
    border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
  }
  .dz-icon svg { width: 22px; height: 22px; stroke: var(--primary); }
  .dz-title { font-size: 15px; font-weight: 600; color: var(--text); }
  .dz-sub { font-size: 13px; color: var(--text-muted); }
  .dz-badge {
    display: inline-block;
    font-size: 11px;
    font-weight: 600;
    letter-spacing: 0.4px;
    padding: 3px 8px;
    border-radius: 4px;
    background: #e0e7ef;
    color: var(--primary);
    margin-top: 4px;
  }
  .dz-file { display: none; align-items: center; gap: 12px; }
  .dz-file.show { display: flex; }
  .dz-file-icon {
    width: 40px; height: 40px; flex-shrink: 0;
    background: var(--primary);
    border-radius: 8px;
    display: flex; align-items: center; justify-content: center;
  }
  .dz-file-icon svg { width: 20px; height: 20px; stroke: #fff; }
  .dz-file-info { flex: 1; text-align: left; min-width: 0; }
  .dz-file-name {
    font-size: 14px; font-weight: 600;
    color: var(--text); white-space: nowrap;
    overflow: hidden; text-overflow: ellipsis;
  }
  .dz-file-size { font-size: 12px; color: var(--text-muted); margin-top: 2px; }
  .dz-remove {
    background: none; border: none;
    cursor: pointer; padding: 6px;
    border-radius: 6px;
    color: var(--text-muted);
    transition: background 0.1s, color 0.1s;
    z-index: 1;
    position: relative;
    flex-shrink: 0;
  }
  .dz-remove:hover { background: #fee2e2; color: #dc2626; }
  .dz-remove svg { width: 16px; height: 16px; stroke: currentColor; display: block; }

  /* ── Convert button ── */
  .btn-convert {
    margin-top: 16px;
    width: 100%;
    padding: 13px 20px;
    background: var(--primary);
    color: #fff;
    font-size: 15px;
    font-weight: 600;
    border: none;
    border-radius: var(--radius-sm);
    cursor: pointer;
    transition: background 0.15s, opacity 0.15s, transform 0.1s;
    display: flex; align-items: center; justify-content: center; gap: 8px;
  }
  .btn-convert:hover:not(:disabled) { background: var(--primary-hover); }
  .btn-convert:active:not(:disabled) { transform: scale(0.99); }
  .btn-convert:disabled { opacity: 0.45; cursor: not-allowed; }
  .btn-convert svg { width: 18px; height: 18px; stroke: #fff; flex-shrink: 0; }

  /* ── Spinner ── */
  .spinner {
    width: 16px; height: 16px;
    border: 2px solid rgba(255,255,255,0.35);
    border-top-color: #fff;
    border-radius: 50%;
    animation: spin 0.65s linear infinite;
    flex-shrink: 0;
  }
  @keyframes spin { to { transform: rotate(360deg); } }

  /* ── Status banner ── */
  .status {
    display: none;
    margin-top: 16px;
    padding: 14px 16px;
    border-radius: var(--radius-sm);
    font-size: 13.5px;
    line-height: 1.55;
    border: 1px solid transparent;
  }
  .status.show { display: block; }
  .status.ok {
    background: var(--success-bg);
    color: var(--success);
    border-color: var(--success-border);
  }
  .status.err {
    background: var(--error-bg);
    color: var(--error);
    border-color: var(--error-border);
  }
  .status.loading {
    background: var(--primary-light);
    color: var(--primary);
    border-color: #bfdbfe;
    display: flex; align-items: center; gap: 10px;
  }
  .status.loading .spinner {
    border-color: rgba(26,79,138,0.25);
    border-top-color: var(--primary);
  }

  /* ── Summary pills ── */
  .summary-pills {
    display: flex; flex-wrap: wrap; gap: 8px;
    margin-top: 10px;
  }
  .pill {
    display: flex; align-items: center; gap: 5px;
    background: #d1fae5;
    color: #065f46;
    border-radius: 20px;
    padding: 4px 12px;
    font-size: 12.5px;
    font-weight: 600;
  }
  .pill svg { width: 13px; height: 13px; stroke: currentColor; }

  /* ── Divider ── */
  .divider { height: 1px; background: var(--border); margin: 0; }

  /* ── Info section ── */
  .info { padding: 20px 28px 24px; background: var(--surface-2); }
  .info-title {
    font-size: 12px;
    font-weight: 700;
    letter-spacing: 0.6px;
    text-transform: uppercase;
    color: var(--text-soft);
    margin-bottom: 12px;
  }
  .info-list { list-style: none; display: flex; flex-direction: column; gap: 8px; }
  .info-list li {
    display: flex; align-items: flex-start; gap: 10px;
    font-size: 13.5px; color: var(--text-muted);
  }
  .info-list li svg {
    width: 15px; height: 15px;
    stroke: var(--primary); flex-shrink: 0; margin-top: 1px;
  }
  code {
    font-family: 'SFMono-Regular', Consolas, monospace;
    font-size: 12px;
    background: #e5e7eb;
    padding: 1px 5px;
    border-radius: 3px;
    color: #374151;
  }
</style>
</head>
<body>
<div class="page">

  <div class="header">
    <div class="header-icon">
      <svg viewBox="0 0 24 24" fill="none" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
        <path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/>
        <polyline points="14 2 14 8 20 8"/>
        <line x1="16" y1="13" x2="8" y2="13"/>
        <line x1="16" y1="17" x2="8" y2="17"/>
        <polyline points="10 9 9 9 8 9"/>
      </svg>
    </div>
    <div class="header-text">
      <h1>CI Benchmarking → CODIFY</h1>
      <p>Convert benchmarking workbooks into codified comparison files</p>
    </div>
  </div>

  <div class="card">
    <div class="card-body">

      <!-- Drop zone -->
      <div id="drop" class="dropzone" role="button" aria-label="Upload Excel file">
        <input type="file" id="file" accept=".xlsx,.xlsm" aria-label="Choose Excel file">

        <div class="dz-idle" id="dz-idle">
          <div class="dz-icon">
            <svg viewBox="0 0 24 24" fill="none" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
              <path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/>
              <polyline points="17 8 12 3 7 8"/>
              <line x1="12" y1="3" x2="12" y2="15"/>
            </svg>
          </div>
          <div>
            <div class="dz-title">Drop your workbook here</div>
            <div class="dz-sub">or click to browse files</div>
          </div>
          <span class="dz-badge">.xlsx &nbsp;/&nbsp; .xlsm</span>
        </div>

        <div class="dz-file" id="dz-file">
          <div class="dz-file-icon">
            <svg viewBox="0 0 24 24" fill="none" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
              <path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/>
              <polyline points="14 2 14 8 20 8"/>
            </svg>
          </div>
          <div class="dz-file-info">
            <div class="dz-file-name" id="dz-file-name">—</div>
            <div class="dz-file-size" id="dz-file-size">—</div>
          </div>
          <button class="dz-remove" id="dz-remove" type="button" aria-label="Remove file" title="Remove file">
            <svg viewBox="0 0 24 24" fill="none" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
              <line x1="18" y1="6" x2="6" y2="18"/>
              <line x1="6" y1="6" x2="18" y2="18"/>
            </svg>
          </button>
        </div>
      </div>

      <!-- Convert button -->
      <button id="convert" class="btn-convert" disabled type="button">
        <svg id="btn-icon" viewBox="0 0 24 24" fill="none" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
          <polyline points="16 16 12 12 8 16"/>
          <line x1="12" y1="12" x2="12" y2="21"/>
          <path d="M20.39 18.39A5 5 0 0 0 18 9h-1.26A8 8 0 1 0 3 16.3"/>
        </svg>
        <span id="btn-label">Convert to CODIFY</span>
      </button>

      <!-- Status -->
      <div id="status" class="status" role="status" aria-live="polite"></div>
    </div>

    <div class="divider"></div>

    <div class="info">
      <div class="info-title">What this tool does</div>
      <ul class="info-list">
        <li>
          <svg viewBox="0 0 24 24" fill="none" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"/></svg>
          Auto-detects companies, products, and CI conditions from the source workbook
        </li>
        <li>
          <svg viewBox="0 0 24 24" fill="none" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"/></svg>
          Builds the full CODIFY comparison structure (38 sheets)
        </li>
        <li>
          <svg viewBox="0 0 24 24" fill="none" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"/></svg>
          Preserves free-text values in <code>raw_text</code> for full traceability
        </li>
        <li>
          <svg viewBox="0 0 24 24" fill="none" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"/></svg>
          Output saved as <code>&lt;filename&gt;_Codified.xlsx</code>
        </li>
      </ul>
    </div>
  </div>

</div>

<script>
const dropEl    = document.getElementById('drop');
const fileInput = document.getElementById('file');
const dzIdle    = document.getElementById('dz-idle');
const dzFile    = document.getElementById('dz-file');
const dzName    = document.getElementById('dz-file-name');
const dzSize    = document.getElementById('dz-file-size');
const dzRemove  = document.getElementById('dz-remove');
const convertBtn= document.getElementById('convert');
const btnLabel  = document.getElementById('btn-label');
const btnIcon   = document.getElementById('btn-icon');
const statusBox = document.getElementById('status');
let selectedFile = null;

function fmtSize(b) {
  if (b < 1024) return b + ' B';
  if (b < 1024*1024) return (b/1024).toFixed(1) + ' KB';
  return (b/(1024*1024)).toFixed(1) + ' MB';
}

function setFile(f) {
  selectedFile = f;
  if (f) {
    dzName.textContent = f.name;
    dzSize.textContent = fmtSize(f.size);
    dzIdle.style.display = 'none';
    dzFile.classList.add('show');
    dropEl.classList.add('has-file');
    convertBtn.disabled = false;
    clearStatus();
  } else {
    dzIdle.style.display = '';
    dzFile.classList.remove('show');
    dropEl.classList.remove('has-file');
    convertBtn.disabled = true;
    fileInput.value = '';
    clearStatus();
  }
}

function clearStatus() {
  statusBox.className = 'status';
  statusBox.innerHTML = '';
}

function showLoading() {
  statusBox.innerHTML = '<div class="spinner"></div><span>Processing your workbook…</span>';
  statusBox.className = 'status loading show';
}

function showSuccess(meta) {
  let html = '<strong>Conversion complete — download started.</strong>';
  if (meta && meta.companies !== undefined) {
    html += '<div class="summary-pills">'
      + pill('building', meta.companies + ' ' + (meta.companies === 1 ? 'company' : 'companies'))
      + pill('package', meta.products + ' ' + (meta.products === 1 ? 'product' : 'products'))
      + pill('activity', meta.conditions + ' CI conditions')
      + pill('layers', meta.sheets + ' sheets')
      + '</div>';
  }
  statusBox.innerHTML = html;
  statusBox.className = 'status ok show';
}

function pill(icon, text) {
  const icons = {
    building: '<svg viewBox="0 0 24 24" fill="none" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect x="2" y="3" width="20" height="18" rx="2"/><path d="M9 3v18M15 3v18M2 9h20M2 15h20"/></svg>',
    package:  '<svg viewBox="0 0 24 24" fill="none" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 16V8a2 2 0 0 0-1-1.73l-7-4a2 2 0 0 0-2 0l-7 4A2 2 0 0 0 3 8v8a2 2 0 0 0 1 1.73l7 4a2 2 0 0 0 2 0l7-4A2 2 0 0 0 21 16z"/></svg>',
    activity: '<svg viewBox="0 0 24 24" fill="none" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="22 12 18 12 15 21 9 3 6 12 2 12"/></svg>',
    layers:   '<svg viewBox="0 0 24 24" fill="none" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polygon points="12 2 2 7 12 12 22 7 12 2"/><polyline points="2 17 12 22 22 17"/><polyline points="2 12 12 17 22 12"/></svg>',
  };
  return '<span class="pill">' + (icons[icon] || '') + text + '</span>';
}

function showError(msg) {
  statusBox.innerHTML = '<strong>Error:</strong> ' + msg;
  statusBox.className = 'status err show';
}

function setBtnLoading(loading) {
  convertBtn.disabled = loading;
  if (loading) {
    btnIcon.outerHTML = '<div class="spinner" id="btn-icon"></div>';
    btnLabel.textContent = 'Converting…';
  } else {
    document.getElementById('btn-icon').outerHTML = '<svg id="btn-icon" viewBox="0 0 24 24" fill="none" stroke="#fff" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="16 16 12 12 8 16"/><line x1="12" y1="12" x2="12" y2="21"/><path d="M20.39 18.39A5 5 0 0 0 18 9h-1.26A8 8 0 1 0 3 16.3"/></svg>';
    btnLabel.textContent = 'Convert to CODIFY';
  }
}

dzRemove.addEventListener('click', (e) => { e.stopPropagation(); setFile(null); });
fileInput.addEventListener('change', (e) => { const f = e.target.files[0]; if (f) setFile(f); });

['dragenter', 'dragover'].forEach(ev =>
  dropEl.addEventListener(ev, (e) => { e.preventDefault(); if (!selectedFile) dropEl.classList.add('over'); })
);
['dragleave', 'drop'].forEach(ev =>
  dropEl.addEventListener(ev, (e) => { e.preventDefault(); dropEl.classList.remove('over'); })
);
dropEl.addEventListener('drop', (e) => {
  const f = e.dataTransfer.files[0];
  if (f) setFile(f);
});

convertBtn.addEventListener('click', async () => {
  if (!selectedFile) return;
  setBtnLoading(true);
  showLoading();

  const fd = new FormData();
  fd.append('file', selectedFile);

  try {
    const r = await fetch('/convert', { method: 'POST', body: fd });
    if (!r.ok) {
      const err = await r.json().catch(() => ({ error: 'Server error ' + r.status }));
      showError(err.error || 'Conversion failed. Please check your workbook and try again.');
      setBtnLoading(false);
      return;
    }
    const meta = JSON.parse(r.headers.get('X-Codify-Summary') || 'null');
    const blob = await r.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    a.download = selectedFile.name.replace(/\.(xlsx|xlsm)$/i, '') + '_Codified.xlsx';
    document.body.appendChild(a);
    a.click();
    a.remove();
    URL.revokeObjectURL(url);
    showSuccess(meta);
  } catch (err) {
    showError('Network error: ' + err.message);
  } finally {
    setBtnLoading(false);
  }
});
</script>
</body>
</html>
"""


@app.route('/')
def index():
    return render_template_string(INDEX_HTML)


@app.route('/convert', methods=['POST'])
def convert():
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'error': 'No file uploaded.'}), 400
    if not f.filename.lower().endswith(('.xlsx', '.xlsm')):
        return jsonify({'error': 'Please upload an .xlsx or .xlsm file.'}), 400

    try:
        in_bytes = io.BytesIO(f.read())
        wb_in = load_workbook(in_bytes, data_only=True)
    except Exception as e:
        return jsonify({'error': f'Could not read workbook: {e}'}), 400

    try:
        products, cond_data = extract_products(wb_in)
        if not products:
            return jsonify({'error': 'No products were detected. Check the Feature Benchmarking sheet.'}), 400
        wb_out = build_workbook(products, cond_data)
    except Exception as e:
        tb = traceback.format_exc()
        app.logger.error(tb)
        return jsonify({'error': f'Conversion failed: {e}'}), 500

    out_buf = io.BytesIO()
    wb_out.save(out_buf)
    out_buf.seek(0)

    summary = {
        'companies': len({p['company_name'] for p in products}),
        'products': len(products),
        'conditions': len(cond_data['conditions']) if cond_data else 0,
        'sheets': len(wb_out.sheetnames),
    }
    import json
    download_name = re.sub(r'\.(xlsx|xlsm)$', '', f.filename, flags=re.IGNORECASE) + '_Codified.xlsx'

    response = send_file(
        out_buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=download_name,
    )
    response.headers['X-Codify-Summary'] = json.dumps(summary)
    response.headers['Access-Control-Expose-Headers'] = 'X-Codify-Summary'
    return response


@app.route('/health')
def health():
    return jsonify({'status': 'ok', 'timestamp': datetime.utcnow().isoformat()})


if __name__ == '__main__':
    print('━' * 50)
    print('  CI Benchmarking → CODIFY Converter')
    print('  Open: http://127.0.0.1:5000')
    print('  Press Ctrl+C to stop')
    print('━' * 50)
    app.run(host='127.0.0.1', port=5000, debug=False)
