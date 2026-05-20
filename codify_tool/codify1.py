#!/usr/bin/env python3
"""
codify.py - Convert a CI Benchmarking workbook into the CODIFY COMPARISON structure.

USAGE
    python codify.py <input.xlsx> [<output.xlsx>]

If output path is omitted, writes "<input_stem>_Codified.xlsx" next to the input.

WHAT IT DOES
    1. Auto-detects the Feature Benchmarking sheet (companies row, products row, feature rows).
    2. Auto-detects the Condition Benchmarking sheet (CI conditions per product).
    3. Builds a clean CODIFY workbook with:
       - 00_Index (summary)
       - 01_Company_Master, 02_Product_Master, 03_Variant_Master
       - 04 Entry Age, 05 Policy Term, 06 Coverage
       - 07A-07E Waiting Periods
       - 08-12 Benefit Amount / Hospital Stay / Area / Payment / Family
       - 13-27 IPD / OPD benefit placeholders
       - 28 Critical Illness Benefits (codified)
       - 28A CI Conditions Coverage (long-format, if conditions sheet found)
       - 29 Misc & Extra Benefits (multi-row)
       - 30 Deductible, 31 Copayment, 32 No Claim Bonus
    4. Source free text retained in a `raw_text` column for traceability.

REQUIREMENTS
    pip install openpyxl
"""

from __future__ import annotations

import re
import sys
import argparse
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# =====================================================================
# SCHEMA (hardcoded, mirrors CODIFY_COMPARISON v2.9)
# =====================================================================
# Each entry: (sheet_name, [list_of_column_headers])
# First 4 columns of every data sheet are: Variant ID, Variant Name, Product Name, Company Name
ID_COLS = ['Variant ID', 'Variant Name', 'Product Name', 'Company Name']

SCHEMA = {
    '01_Company_Master': ['Company ID', 'Company Name'],
    '02_Product_Master': ['Product ID', 'Product Name'],
    '03_Variant_Master': ['Variant ID', 'Variant Name', 'Product Name', 'Company Name', 'Plan Type'],
    '04_Entry_Age': ID_COLS + [
        'min_entry_age_value', 'min_entry_age_unit',
        'min_entry_age_in_days', 'min_entry_age_in_months', 'min_entry_age_in_years',
        'min_child_age_value', 'min_child_age_unit',
        'min_child_age_in_days', 'min_child_age_in_months', 'min_child_age_in_years',
        'max_entry_age_value', 'max_entry_age_unit',
        'max_entry_age_in_days', 'max_entry_age_in_months', 'max_entry_age_in_years',
        'max_child_age_value', 'max_child_age_unit',
        'max_child_age_in_days', 'max_child_age_in_months', 'max_child_age_in_years',
        'max_renewal_age_value', 'max_renewal_age_unit',
        'max_renewal_age_in_days', 'max_renewal_age_in_months', 'max_renewal_age_in_years',
        'special_conditions', 'raw_text',
    ],
    '05_Policy_Term': ID_COLS + [
        'policy_term_value', 'policy_term_unit',
        'policy_term_in_days', 'policy_term_in_months', 'policy_term_in_years',
        'special_conditions', 'raw_text',
    ],
    '06_Coverage': ID_COLS + [
        'coverage_max_age_value', 'coverage_max_age_unit',
        'coverage_max_age_in_days', 'coverage_max_age_in_months', 'coverage_max_age_in_years',
        'renewal_max_age_value', 'renewal_max_age_unit',
        'renewal_max_age_in_days', 'renewal_max_age_in_months', 'renewal_max_age_in_years',
        'coverage_linked_to_base_plan', 'requires_continuous_renewal',
        'coverage_termination_condition', 'lifetime_renewal_eligibility',
        'lifetime_renewal_condition', 'special_conditions', 'raw_text',
    ],
    '07A_Waiting_Period_Standard': ID_COLS + [
        'waiting_period_type', 'waiting_period_value', 'waiting_period_unit',
        'waiting_period_in_days', 'waiting_period_in_months', 'waiting_period_in_years',
        'waiting_period_conditions', 'special_conditions', 'raw_text',
    ],
    '07B_Waiting_Period_Accident': None,  # same as 07A
    '07C_Waiting_Period_Specific': None,
    '07D_Waiting_Period_Special(90D)': None,
    '07E_Waiting_Period_Other(12MWP)': None,
    '08_Policy_Year_Benefit_Amt': ID_COLS + [
        'policy_year_max_benefit', 'policy_year_max_benefit_currency',
        'si_applicability', 'special_conditions', 'raw_text',
    ],
    '09_Hospital_Stay_Benefit_Amt': ID_COLS + [
        'per_hospital_stay_max_benefit', 'per_hospital_stay_max_benefit_currency',
        'si_applicability', 'special_conditions', 'raw_text',
    ],
    '10_Area_Covered': ID_COLS + [
        'coverage_area_type', 'exclusion_coverage_area', 'special_conditions', 'raw_text',
    ],
    '11_Payment_Frequency': ID_COLS + [
        'annual_mode', 'semi_annual_mode', 'quarterly_mode', 'monthly_mode',
        'mode_of_payment', 'payment_requires_base_plan', 'special_conditions', 'raw_text',
    ],
    '12_Family_Discounts': ID_COLS + [
        'has_family_discount', 'family_discount_type', 'family_discount_percentage',
        'discount_basis', 'special_conditions', 'raw_text',
    ],
    # IPD / OPD benefit placeholders — schema-only, source is CI-focused
    '13_Room_Board_Inpatient': ID_COLS + [
        'benefit_name', 'room_type_applicablity', 'room_type', 'room_rent_amount',
        'room_rent_amount_unit', 'room_rent_max_days', 'room_rent_payable_mode',
        'si_applicability', 'special_conditions', 'raw_text',
    ],
    '14_ICU_Room_Inpatient': ID_COLS + [
        'benefit_name', 'room_type_applicablity', 'room_type',
        'icu_room_rent_amount', 'icu_room_rent_amount_unit', 'icu_room_rent_max_days',
        'icu_rent_type', 'icu_room_rent_payable_mode', 'si_applicability',
        'special_conditions', 'raw_text',
        'icu_rent_percentage_of_si', 'icu_layer_special_conditions', 'icu_rent_multiplier',
        'icu_limit_type', 'icu_cap_numeric', 'icu_cap_numeric_payable_mode',
        'extracted_multiplier', 'standardized_icu_days', 'effective_icu_daily_coverage',
        'icu_daily_gap', 'icu_financial_risk_category', 'icu_duration_risk_bucket',
        'icu_pricing_strategy',
    ],
    '15_Medical_Services_Fees': ID_COLS + [
        'benefit_name', 'benefit_level', 'sub_benefit_name',
        'medical_services_payable_mode', 'medical_services_amount_unit',
        'medical_services_amount_limit', 'medical_services_max_duration_value',
        'medical_services_max_duration_unit', 'special_conditions', 'raw_text',
        'take_home_limit_structure', 'take_home_limit_amount', 'take_home_limit_days',
        'limit_structure', 'medical_services_amount_limit_normalized',
        'standardized_category', 'take_home_meds_risk_band',
        'sub_benefit_count_per_variant', 'ease_of_claim_score', 'claim_friction_band',
        'inflation_resilience_type', 'benefit_structure_type',
    ],
    '16_IPD_Physician_Fees': ID_COLS + [
        'benefit_name', 'physician_fees_applicability', 'physician_fees_amount',
        'physician_fees_amount_unit', 'physician_fees_max_limit',
        'physician_fees_max_limit_unit', 'physician_fees_payable_mode',
        'si_applicability', 'special_conditions', 'raw_text',
        'fee_structure_type', 'physician_fees_amount_numeric',
        'physician_fees_max_limit_normalized', 'physician_fees_unit_normalized',
        'doctor_fee_gap_thb', 'doctor_fee_gap_risk_band', 'doctor_fee_adequacy_flag',
    ],
    '17_Medical_Expenses_Surgery_Pro': ID_COLS + [
        'benefit_name', 'benefit_level', 'sub_benefit_name', 'amount_limit',
        'amount_unit', 'payable_mode', 'max_duration_value', 'max_duration_unit',
        'si_applicability', 'special_conditions', 'raw_text',
    ],
    '18_Day_Surgery_Major': ID_COLS + [
        'benefit_name', 'day_surgery_applicability', 'day_surgery_amount',
        'day_surgery_amount_unit', 'day_surgery_max_limit', 'day_surgery_max_limit_unit',
        'day_surgery_payable_mode', 'si_applicability', 'special_conditions', 'raw_text',
    ],
    '19_Diagnostic_Services_IPD': ID_COLS + [
        'benefit_name', 'benefit_level', 'sub_benefit_name', 'amount_limit',
        'amount_unit', 'payable_mode', 'max_duration_value', 'max_duration_unit',
        'si_applicability', 'special_conditions', 'raw_text',
    ],
    '20_OPD_Accident_24_Hours': ID_COLS + [
        'benefit_name', 'opd_accident_applicability', 'opd_accident_amount_value',
        'opd_accident_amount_unit', 'opd_accident_payable_mode',
        'opd_accident_max_limit', 'opd_accident_max_limit_unit',
        'si_applicability', 'special_conditions', 'raw_text',
    ],
    '21_Rehabilitation_Expenses_IPD': ID_COLS + [
        'benefit_name', 'rehab_applicability', 'rehab_amount_value', 'rehab_amount_unit',
        'rehab_payable_mode', 'rehab_max_limit_value', 'rehab_max_limit_unit',
        'si_applicability', 'special_conditions', 'raw_text',
    ],
    '22_Dialysis_Treatment': ID_COLS + [
        'benefit_name', 'dialysis_applicability', 'dialysis_type',
        'dialysis_amount_value', 'dialysis_amount_unit', 'dialysis_payable_mode',
        'dialysis_max_limit_value', 'dialysis_max_limit_unit',
        'si_applicability', 'special_conditions', 'raw_text',
    ],
    '23_Cancer_Treatment_Radiation': ID_COLS + [
        'benefit_name', 'cancer_radiation_applicability', 'treatment_type',
        'cancer_radiation_amount_value', 'cancer_radiation_amount_unit',
        'cancer_radiation_payable_mode', 'cancer_radiation_max_limit_value',
        'cancer_radiation_max_limit_unit', 'si_applicability', 'special_conditions', 'raw_text',
    ],
    '24_Chemo_Treatment_Expenses': ID_COLS + [
        'benefit_name', 'chemo_applicability', 'chemotherapy_type',
        'chemo_amount_value', 'chemo_amount_unit', 'chemo_payable_mode',
        'chemo_max_limit_value', 'chemo_max_limit_unit',
        'si_applicability', 'special_conditions', 'raw_text',
    ],
    '25_Emergency_Ambulance_Services': ID_COLS + [
        'benefit_name', 'ambulance_applicability', 'ambulance_amount_value',
        'ambulance_amount_unit', 'ambulance_payable_mode',
        'ambulance_max_limit_value', 'ambulance_max_limit_unit',
        'si_applicability', 'special_conditions', 'raw_text',
    ],
    '26_Minor_Surgery_Expenses': ID_COLS + [
        'benefit_name', 'minor_surgery_applicability', 'minor_surgery_amount_value',
        'minor_surgery_amount_unit', 'minor_surgery_payable_mode',
        'minor_surgery_max_limit_value', 'minor_surgery_max_limit_unit',
        'si_applicability', 'special_conditions', 'raw_text',
    ],
    '27_OPD_Medical_Services': ID_COLS + [
        'benefit_name', 'opd_applicability', 'opd_treatment_type',
        'opd_amount_value', 'opd_amount_unit', 'opd_payable_mode',
        'opd_max_limit _value', 'opd_max_limit_unit',
        'si_applicability', 'special_conditions', 'raw_text',
    ],
    '28_Critical_Illness_Benefits': ID_COLS + [
        'benefit_name', 'ci_benefit_type', 'ci_trigger_condition', 'ci_amount_value',
        'ci_amount_type', 'ci_payable_mode', 'si_applicability', 'special_conditions',
        'raw_text', 'ci_covered_flag', 'ci_missing_reason', 'ci_amount_basis',
        'ci_amount_numeric', 'ci_amount_multiplier', 'ci_amount_currency',
        'ci_illness_scope', 'ci_scope_details', 'ci_max_claims_count',
        'ci_max_claims_period', 'ci_trigger_stage', 'ci_requires_hospitalisation_flag',
        'ci_requires_waiting_period_flag', 'ci_is_over_and_above_flag',
        'ci_reduces_core_limit_flag', 'ci_multiple_years_flag', 'ci_payable_mode_std',
        'ci_archetype', 'ci_generosity_score', 'ci_generosity_reason',
        'ci_present_flag', 'ci_archetype_group', 'ci_above_si_numeric',
    ],
    '28A_CI_Conditions_Coverage': ID_COLS + [
        'condition_sno', 'condition_category', 'condition_name',
        'condition_severity', 'condition_stage', 'coverage_status',
    ],
    '29_Misc_&_Extra_Benefits': ID_COLS + [
        'benefit_name', 'sub_benefit_name', 'applicability', 'payable_mode',
        'amount_value', 'amount_unit', 'max_limit_value', 'max_limit_unit',
        'si_applicability', 'special_conditions', 'raw_text',
    ],
    '30_Deductible': ID_COLS + [
        'benefit_name', 'deductible_applicability', 'deductible_type',
        'deductible_amount_value', 'deductible_amount_unit', 'deductible_frequency',
        'special_conditions', 'raw_text',
    ],
    '31_Copayment': ID_COLS + [
        'benefit_name', 'copayment_applicability', 'copayment_type',
        'copayment_value', 'copayment_unit', 'copayment_trigger',
        'special_conditions', 'raw_text',
    ],
    '32_No_Claim_Bonus': ID_COLS + [
        'benefit_name', 'ncb_type', 'ncb_rate_value', 'ncb_rate_unit',
        'ncb_max_limit', 'ncb_max_limit_unit', 'ncb_reset_condition',
        'special_conditions', 'raw_text',
    ],
}
# Fill waiting period clones
for k in ('07B_Waiting_Period_Accident', '07C_Waiting_Period_Specific',
         '07D_Waiting_Period_Special(90D)', '07E_Waiting_Period_Other(12MWP)'):
    SCHEMA[k] = SCHEMA['07A_Waiting_Period_Standard']


# =====================================================================
# COMPANY SHORT NAME REGISTRY (extensible)
# Maps long legal names to short display names. Falls back to the long
# name if no match — auto-detection handles the rest.
# =====================================================================
COMPANY_SHORT_MAP = {
    'aia philippines life and general insurance company, inc.': 'AIA Philippines',
    'bpi aia life assurance corporation': 'BPI AIA Life Assurance',
    'sun life of canada (philippines) incorporated': 'Sun Life Philippines',
    'the insular life assurance company, ltd': 'Insular Life',
    'pru life insurance corporation of u.k.': 'Pru Life UK',
    'fwd life insurance corporation': 'FWD Life Philippines',
    'the manufacture life insurance co. (phis.), inc.': 'Manulife Philippines',
    'allianz pnb insurance, inc.': 'Allianz PNB',
    'axa philippines life and general insurance corporation': 'AXA Philippines',
    'singlife philippines life insurance corporation': 'Singlife Philippines',
    # Thai entries (in case the script is used on Thai files too)
    'prudential life assurance public company limited (thailand)': 'Prudential Thailand',
    'thai life insurance public company limited': 'Thai Life Insurance',
    'fwd life insurance public company limited': 'FWD Life Thailand',
    'aia thailand': 'AIA Thailand',
    'allianz ayudhya assurance public company limited': 'Allianz Ayudhya',
    'krungthai-axa life insurance public company limited': 'Krungthai-AXA',
}


def short_name(long_name: str) -> str:
    """Best-effort short company name lookup; falls back to a trimmed long name."""
    key = (long_name or '').strip().lower()
    if key in COMPANY_SHORT_MAP:
        return COMPANY_SHORT_MAP[key]
    # Heuristic: take first 2-3 words before commas/Inc./Ltd
    s = long_name.strip()
    s = re.split(r',|\(|inc\.|incorporated|corporation|limited|ltd', s, flags=re.IGNORECASE)[0]
    return s.strip() or long_name


# =====================================================================
# AUTO-DETECTION
# =====================================================================
def is_na(v) -> bool:
    if v is None:
        return True
    s = str(v).strip()
    return s == '' or s.upper() in {'NA', 'N/A', 'NONE', '-'}


def clean(v) -> Optional[str]:
    if is_na(v):
        return None
    return str(v).strip()


def detect_feature_sheet(wb) -> Optional[tuple]:
    """
    Find the Feature Benchmarking sheet by signature:
    - A row labeled 'Company' (or 'Companies') in early rows
    - A row labeled 'Products' (or 'Product') below it
    - Companies/products spread across columns 4+

    Returns (worksheet, company_row, product_row, feature_start_row, first_data_col, last_data_col)
    or None if not found.
    """
    candidates = []
    for sn in wb.sheetnames:
        if 'condition' in sn.lower():
            continue  # skip the conditions sheet
        ws = wb[sn]
        company_row = None
        product_row = None
        label_col = None  # column that contains the "Company"/"Products" labels
        # Search first 8 rows for labels in cols A-C
        for r in range(1, min(9, ws.max_row + 1)):
            for c in range(1, 5):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str):
                    vs = v.strip().lower()
                    if vs in {'company', 'companies', 'insurance company'} and company_row is None:
                        company_row = r
                        label_col = c
                    elif vs in {'products', 'product', 'product name'} and product_row is None:
                        product_row = r
                        if label_col is None:
                            label_col = c
        if company_row and product_row:
            # find the data column range — skip past the label column
            first_data_col = None
            last_data_col = None
            start_c = (label_col + 1) if label_col else 2
            for c in range(start_c, ws.max_column + 1):
                v = ws.cell(row=company_row, column=c).value
                if v and not is_na(v):
                    if first_data_col is None:
                        first_data_col = c
                    last_data_col = c
            if first_data_col and last_data_col and last_data_col >= first_data_col:
                # find feature start row: first row after product_row where col B/C has a text feature name
                # and the data column has a value
                feat_start = None
                for r in range(product_row + 1, min(product_row + 10, ws.max_row + 1)):
                    for c in range(1, 4):
                        v = ws.cell(row=r, column=c).value
                        if isinstance(v, str) and v.strip() and v.strip().lower() not in {
                            'features', 'feature', 'benefits', 'benefit', 's.no.', 'sno', 'sl.no.',
                            'individual_product_detail', 'product_brochure', 'category', 'no.'
                        } and not v.lower().startswith('http'):
                            feat_start = r
                            break
                    if feat_start:
                        break
                if feat_start:
                    candidates.append({
                        'sheet': ws,
                        'name': sn,
                        'company_row': company_row,
                        'product_row': product_row,
                        'feature_start_row': feat_start,
                        'first_data_col': first_data_col,
                        'last_data_col': last_data_col,
                    })
    if not candidates:
        return None
    # Prefer the one with most data columns × most feature rows
    best = max(candidates,
               key=lambda c: (c['last_data_col'] - c['first_data_col'] + 1)
                              * (c['sheet'].max_row - c['feature_start_row'] + 1))
    return best


def detect_condition_sheet(wb) -> Optional[dict]:
    """
    Find a Condition Benchmarking sheet. Signature:
    - Sheet name contains 'condition' OR has columns 'Condition Name', 'Major/Minor', 'Stage'
    - Companies and products listed across columns

    Returns dict with sheet, company_row, product_row, condition_start_row, first/last_data_col,
    sno_col, category_col, name_col, severity_col, stage_col
    """
    for sn in wb.sheetnames:
        ws = wb[sn]
        # Quick name check
        name_match = 'condition' in sn.lower()
        # Look for header signature in first 6 rows
        header_row = None
        cols = {}
        for r in range(1, min(7, ws.max_row + 1)):
            for c in range(1, min(15, ws.max_column + 1)):
                v = ws.cell(row=r, column=c).value
                if not isinstance(v, str):
                    continue
                vs = v.strip().lower()
                if 'condition name' in vs:
                    cols['name'] = c
                    header_row = r
                elif vs in {'major/minor', 'severity', 'type'}:
                    cols['severity'] = c
                elif 'condition category' in vs or vs == 'category':
                    if 'category' not in cols:
                        cols['category'] = c
                elif vs in {'stage (early/late)', 'stage'}:
                    cols['stage'] = c
                elif vs in {'s.no.', 'sno', 's.no', 'sl.no.', 'no.'}:
                    cols['sno'] = c
        if header_row is None or 'name' not in cols:
            if not name_match:
                continue
            else:
                # named like condition sheet but no header found - skip
                continue

        # Find company/product rows above the header
        company_row = None
        product_row = None
        for r in range(1, header_row):
            for c in range(1, 6):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str):
                    vs = v.strip().lower()
                    if vs in {'company', 'companies'} and company_row is None:
                        company_row = r
                    elif vs in {'products', 'product'} and product_row is None:
                        product_row = r

        # Find data column range
        first_data_col = max(cols.values()) + 1
        last_data_col = first_data_col
        if company_row:
            for c in range(first_data_col, ws.max_column + 1):
                v = ws.cell(row=company_row, column=c).value
                if v and not is_na(v):
                    last_data_col = c
        elif product_row:
            for c in range(first_data_col, ws.max_column + 1):
                v = ws.cell(row=product_row, column=c).value
                if v and not is_na(v):
                    last_data_col = c

        # Find last condition row
        last_cond_row = header_row
        for r in range(header_row + 1, ws.max_row + 1):
            v = ws.cell(row=r, column=cols['name']).value
            if v and str(v).strip():
                last_cond_row = r

        return {
            'sheet': ws,
            'name': sn,
            'header_row': header_row,
            'company_row': company_row,
            'product_row': product_row,
            'condition_start_row': header_row + 1,
            'last_condition_row': last_cond_row,
            'cols': cols,
            'first_data_col': first_data_col,
            'last_data_col': last_data_col,
        }
    return None


# =====================================================================
# EXTRACTION
# =====================================================================
def extract_products(wb) -> tuple[list[dict], Optional[dict]]:
    """
    Returns (products_list, condition_data) where products_list is:
        [{ 'company_name', 'product_name', 'features': {feature_name: value, ...} }, ...]
    and condition_data may be None if no condition sheet found, else:
        { 'conditions': [...], 'product_cols': [...] }
    """
    feat = detect_feature_sheet(wb)
    if feat is None:
        raise RuntimeError(
            'Could not auto-detect a Feature Benchmarking sheet.\n'
            'Expected a sheet with "Company" and "Products" labels in early rows '
            'and product names spread across columns.'
        )

    ws = feat['sheet']
    company_row = feat['company_row']
    product_row = feat['product_row']
    feat_start = feat['feature_start_row']
    first_c = feat['first_data_col']
    last_c = feat['last_data_col']

    # Detect which columns in the feature area contain the feature name
    # In the Philippines layout, category is col B (2) and feature is col C (3).
    # We auto-detect: scan a few feature rows and find which left-side col is most consistently text.
    cat_col, feat_col = None, None
    # Candidate: the col with the most non-empty strings between feat_start and feat_start+20
    counts = {1: 0, 2: 0, 3: 0}
    for r in range(feat_start, min(feat_start + 30, ws.max_row + 1)):
        for c in (1, 2, 3):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() and not v.strip().isdigit():
                counts[c] += 1
    # Two highest-count columns are the category/feature pair
    sorted_cols = sorted(counts.items(), key=lambda x: -x[1])
    top = [c for c, n in sorted_cols if n > 5]
    if len(top) >= 2:
        cat_col, feat_col = sorted(top[:2])  # left-most = category, right = feature
    elif len(top) == 1:
        feat_col = top[0]
        cat_col = None

    if feat_col is None:
        # fall back to col 3
        feat_col = 3
        cat_col = 2

    # Build product records
    products = []
    for c in range(first_c, last_c + 1):
        company = ws.cell(row=company_row, column=c).value
        product = ws.cell(row=product_row, column=c).value
        if is_na(company) or is_na(product):
            continue
        rec = {
            'company_name': str(company).strip().rstrip('\t').strip(),
            'product_name': str(product).strip().rstrip('\t').strip(),
            'features': {},
        }
        for r in range(feat_start, ws.max_row + 1):
            feat_name = ws.cell(row=r, column=feat_col).value
            if not feat_name or not isinstance(feat_name, str):
                continue
            fn = feat_name.strip()
            if not fn:
                continue
            rec['features'][fn] = ws.cell(row=r, column=c).value
        products.append(rec)

    # Condition data
    cond_info = detect_condition_sheet(wb)
    cond_data = None
    if cond_info:
        ws2 = cond_info['sheet']
        cols = cond_info['cols']
        # Build per-product map from condition sheet's product row
        cond_product_cols = []
        if cond_info['product_row']:
            for c in range(cond_info['first_data_col'], cond_info['last_data_col'] + 1):
                pn = ws2.cell(row=cond_info['product_row'], column=c).value
                if pn and not is_na(pn):
                    cond_product_cols.append({
                        'col': c,
                        'product_name': str(pn).strip().rstrip('\t').strip(),
                    })

        conditions = []
        for r in range(cond_info['condition_start_row'], cond_info['last_condition_row'] + 1):
            name = ws2.cell(row=r, column=cols['name']).value
            if not name or not str(name).strip():
                continue
            cond_row = {
                'sno': ws2.cell(row=r, column=cols.get('sno', 1)).value if 'sno' in cols else r - cond_info['condition_start_row'] + 1,
                'category': ws2.cell(row=r, column=cols['category']).value if 'category' in cols else None,
                'name': str(name).strip(),
                'severity': ws2.cell(row=r, column=cols['severity']).value if 'severity' in cols else None,
                'stage': ws2.cell(row=r, column=cols['stage']).value if 'stage' in cols else None,
                'coverage': {},
            }
            for cp in cond_product_cols:
                cond_row['coverage'][cp['product_name']] = ws2.cell(row=r, column=cp['col']).value
            conditions.append(cond_row)

        cond_data = {'conditions': conditions, 'product_cols': cond_product_cols}

    return products, cond_data


# =====================================================================
# PARSING / CODIFICATION HELPERS
# =====================================================================
def parse_age(text):
    if is_na(text):
        return (None, None)
    s = str(text).strip().lower()
    m = re.search(r'(\d+)\s*day', s)
    if m: return (int(m.group(1)), 'Day')
    m = re.search(r'(\d+)\s*month', s)
    if m: return (int(m.group(1)), 'Month')
    m = re.search(r'(\d+)\s*year', s)
    if m: return (int(m.group(1)), 'Year')
    m = re.search(r'up to\s*(\d+)', s)
    if m: return (int(m.group(1)), 'Year')
    m = re.search(r'^(\d+)$', s)
    if m: return (int(m.group(1)), 'Year')
    return (None, None)


def parse_policy_term(text):
    if is_na(text): return (None, None)
    s = str(text).strip().lower()
    if 'whole life' in s or 'lifetime' in s or 'until maturity' in s:
        return ('Whole Life', None)
    if 'monthly renew' in s:
        return (1, 'Month')
    m = re.search(r'until age\s*(\d+)', s)
    if m: return (int(m.group(1)), 'Year (to age)')
    m = re.search(r'up to age\s*(\d+)', s)
    if m: return (int(m.group(1)), 'Year (to age)')
    m = re.search(r'(\d+)\s*year', s)
    if m: return (int(m.group(1)), 'Year')
    return (None, None)


def parse_amount(text):
    if is_na(text): return (None, None, None)
    s = str(text).strip()
    m = re.search(r'(?:php|p|₱|thb|baht|usd|\$)\s*([\d,]+(?:\.\d+)?)', s, re.IGNORECASE)
    if m:
        val = float(m.group(1).replace(',', ''))
        ccy_match = re.search(r'(php|thb|usd)', s, re.IGNORECASE)
        ccy = ccy_match.group(1).upper() if ccy_match else None
        if not ccy:
            if '₱' in s: ccy = 'PHP'
            elif '$' in s: ccy = 'USD'
        return (val, ccy, s)
    m = re.search(r'^([\d,]+(?:\.\d+)?)$', s)
    if m:
        return (float(m.group(1).replace(',', '')), None, s)
    return (None, None, s)


def parse_pct(text):
    if is_na(text): return None
    s = str(text).strip()
    m = re.search(r'(\d+(?:\.\d+)?)\s*%', s)
    if m: return float(m.group(1)) / 100.0
    return None


def parse_waiting_period(text):
    if is_na(text): return (None, None)
    s = str(text).strip().lower()
    m = re.search(r'(\d+)\s*day', s)
    if m: return (int(m.group(1)), 'Day')
    m = re.search(r'(\d+)\s*month', s)
    if m: return (int(m.group(1)), 'Month')
    m = re.search(r'(\d+)\s*year', s)
    if m: return (int(m.group(1)), 'Year')
    return (None, None)


def feat(features: dict, *keys, default=None):
    """Get the first non-NA value for any of the given feature keys (case-insensitive)."""
    lc_map = {k.lower(): k for k in features.keys()}
    for k in keys:
        actual = lc_map.get(k.lower())
        if actual is not None:
            v = features[actual]
            if not is_na(v):
                return v
    return default


def make_sheet_name(number: int, label: str) -> str:
    """Build an Excel-safe sheet name: '33_Wellness_Bonus'.
    Excel limit is 31 characters; we trim aggressively if needed.
    Invalid chars: : \\ / ? * [ ]
    """
    safe = re.sub(r"[:\\/?*\[\]'\"]", '', label)
    safe = re.sub(r'\s+', '_', safe.strip())
    safe = re.sub(r'_+', '_', safe)
    prefix = f"{number:02d}_"
    max_label_len = 31 - len(prefix)
    if len(safe) > max_label_len:
        safe = safe[:max_label_len].rstrip('_')
    return prefix + safe


# =====================================================================
# REGISTRY BUILDING
# =====================================================================
def build_registry(products: list[dict]) -> dict:
    """Assigns C##, P##, V## IDs."""
    unique_companies = []
    for p in products:
        c = p['company_name']
        if c not in unique_companies:
            unique_companies.append(c)
    company_id_by_full = {c: f"C{i+1:02d}" for i, c in enumerate(unique_companies)}

    prod_count = {}
    for p in products:
        full = p['company_name']
        cid = company_id_by_full[full]
        prod_count.setdefault(cid, 0)
        prod_count[cid] += 1
        p['_company_id'] = cid
        p['_company_short'] = short_name(full)
        p['_product_id'] = f"{cid}-P{prod_count[cid]:02d}"
        p['_variant_id'] = f"{p['_product_id']}-V01"
        p['_variant_name'] = p['product_name']

    return {
        'unique_companies': unique_companies,
        'company_id_by_full': company_id_by_full,
    }


# =====================================================================
# WORKBOOK BUILDING
# =====================================================================
HEADER_ROW = 5  # blank rows 1-4, headers on row 5, data row 6+

# =====================================================================
# STYLING (non-negotiables)
#   • All fonts: Arial 10
#   • Headers: bold, 25% grey fill (#BFBFBF), Arial 10
#   • No gridlines anywhere
#   • No borders on any cell
# =====================================================================
ARIAL_10 = Font(name='Arial', size=10)
ARIAL_10_BOLD = Font(name='Arial', size=10, bold=True)
HDR_FILL = PatternFill('solid', start_color='BFBFBF')  # 25% grey
HDR_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
DATA_ALIGN = Alignment(vertical='top', wrap_text=True)
NO_BORDER = Border(left=Side(border_style=None), right=Side(border_style=None),
                   top=Side(border_style=None), bottom=Side(border_style=None))


def init_sheet(wb: Workbook, name: str, headers: list[str], header_row: int = HEADER_ROW) -> Worksheet:
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False  # gridlines off
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.fill = HDR_FILL
        cell.font = ARIAL_10_BOLD
        cell.alignment = HDR_ALIGN
        cell.border = NO_BORDER
    ws.row_dimensions[header_row].height = 36
    # Column widths
    for c, h in enumerate(headers, 1):
        col_letter = get_column_letter(c)
        if c <= 4:
            ws.column_dimensions[col_letter].width = {1: 14, 2: 28, 3: 28, 4: 24}.get(c, 18)
        elif 'raw' in h.lower() or 'special' in h.lower() or 'description' in h.lower() \
                or 'condition_name' in h.lower():
            ws.column_dimensions[col_letter].width = 45
        else:
            ws.column_dimensions[col_letter].width = 18
    ws.freeze_panes = ws.cell(row=header_row + 1, column=5)
    return ws


def style_data_row(ws: Worksheet, row_idx: int, n_cols: int):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = ARIAL_10
        cell.alignment = DATA_ALIGN
        cell.border = NO_BORDER


def write_row(ws: Worksheet, row_idx: int, values: list, n_cols: int):
    for c, v in enumerate(values, 1):
        ws.cell(row=row_idx, column=c, value=v)
    style_data_row(ws, row_idx, n_cols)


def to_dmy(value, unit):
    """Convert (value, unit) to (days, months, years) buckets."""
    if value is None or not isinstance(value, (int, float)):
        return (None, None, None)
    if unit == 'Day': return (value, None, None)
    if unit == 'Month': return (None, value, None)
    if unit in ('Year', 'Year (to age)'): return (None, None, value)
    return (None, None, None)


def build_workbook(products: list[dict], cond_data: Optional[dict]) -> Workbook:
    """Construct the codified workbook in memory."""
    registry = build_registry(products)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # --------- 00_Index (will be populated last but created first for ordering)
    ws_idx = wb.create_sheet('00_Index')

    # --------- 01_Company_Master
    ws = init_sheet(wb, '01_Company_Master', SCHEMA['01_Company_Master'])
    n_cols = len(SCHEMA['01_Company_Master'])
    for i, full in enumerate(registry['unique_companies'], 1):
        cid = registry['company_id_by_full'][full]
        write_row(ws, HEADER_ROW + i, [cid, short_name(full)], n_cols)

    # --------- 02_Product_Master
    ws = init_sheet(wb, '02_Product_Master', SCHEMA['02_Product_Master'])
    n_cols = len(SCHEMA['02_Product_Master'])
    for i, p in enumerate(products, 1):
        write_row(ws, HEADER_ROW + i, [p['_product_id'], p['product_name']], n_cols)

    # --------- 03_Variant_Master
    ws = init_sheet(wb, '03_Variant_Master', SCHEMA['03_Variant_Master'])
    n_cols = len(SCHEMA['03_Variant_Master'])
    for i, p in enumerate(products, 1):
        write_row(ws, HEADER_ROW + i, [
            p['_variant_id'], p['_variant_name'], p['product_name'],
            p['_company_short'], 'Standalone',
        ], n_cols)

    # --------- 04_Entry_Age
    ws = init_sheet(wb, '04_Entry_Age', SCHEMA['04_Entry_Age'])
    n_cols = len(SCHEMA['04_Entry_Age'])
    for i, p in enumerate(products, 1):
        min_raw = feat(p['features'], 'Entry Age — Minimum', 'Entry Age - Minimum',
                       'Entry Age Minimum', 'Min Entry Age')
        max_raw = feat(p['features'], 'Entry Age — Maximum', 'Entry Age - Maximum',
                       'Entry Age Maximum', 'Max Entry Age')
        renewal_raw = feat(p['features'], 'Renewal / Continuity', 'Renewal', 'Continuity')

        min_v, min_u = parse_age(min_raw)
        max_v, max_u = parse_age(max_raw)
        mind, minm, miny = to_dmy(min_v, min_u)
        maxd, maxm, maxy = to_dmy(max_v, max_u)
        ren_max = None
        if not is_na(renewal_raw):
            m = re.search(r'(\d+)', str(renewal_raw))
            if m and 'age' in str(renewal_raw).lower():
                ren_max = int(m.group(1))
        is_child = isinstance(min_v, int) and min_v < 18
        row = [
            p['_variant_id'], p['_variant_name'], p['product_name'], p['_company_short'],
            min_v if min_v is not None else 'NA', min_u or 'NA', mind, minm, miny,
            min_v if is_child else 'NA', min_u if is_child else 'NA',
            mind if is_child else None, minm if is_child else None, miny if is_child else None,
            max_v if max_v is not None else 'NA', max_u or 'NA', maxd, maxm, maxy,
            'NA', 'NA', None, None, None,
            ren_max if ren_max is not None else 'NA',
            'Year' if ren_max is not None else 'NA', None, None,
            ren_max if ren_max is not None else None,
            clean(renewal_raw) or 'NA',
            f"Min: {min_raw or 'NA'} | Max: {max_raw or 'NA'}",
        ]
        write_row(ws, HEADER_ROW + i, row, n_cols)

    # --------- 05_Policy_Term
    ws = init_sheet(wb, '05_Policy_Term', SCHEMA['05_Policy_Term'])
    n_cols = len(SCHEMA['05_Policy_Term'])
    for i, p in enumerate(products, 1):
        pt = feat(p['features'], 'Policy Term')
        val, unit = parse_policy_term(pt)
        days = months = years = None
        if isinstance(val, int):
            if unit == 'Month': months = val
            elif unit in ('Year', 'Year (to age)'): years = val
        row = [
            p['_variant_id'], p['_variant_name'], p['product_name'], p['_company_short'],
            val if val is not None else 'NA', unit or 'NA', days, months, years,
            'NA', clean(pt) or 'NA',
        ]
        write_row(ws, HEADER_ROW + i, row, n_cols)

    # --------- 06_Coverage
    ws = init_sheet(wb, '06_Coverage', SCHEMA['06_Coverage'])
    n_cols = len(SCHEMA['06_Coverage'])
    for i, p in enumerate(products, 1):
        pt = feat(p['features'], 'Policy Term')
        renewal = feat(p['features'], 'Renewal / Continuity', 'Renewal', 'Continuity')
        val, unit = parse_policy_term(pt)
        is_lifetime = isinstance(val, str) and val == 'Whole Life'
        if isinstance(val, int) and unit in ('Year (to age)',):
            cov_max = val
        elif is_lifetime:
            cov_max = 99
        elif isinstance(val, int) and unit == 'Year':
            cov_max = val
        else:
            cov_max = None
        row = [
            p['_variant_id'], p['_variant_name'], p['product_name'], p['_company_short'],
            cov_max if cov_max is not None else 'NA', 'Year' if cov_max is not None else 'NA',
            None, None, cov_max,
            cov_max if cov_max is not None else 'NA', 'Year' if cov_max is not None else 'NA',
            None, None, cov_max,
            'Standalone', 'Yes' if not is_na(renewal) else 'NA',
            clean(renewal) or 'NA', 'Yes' if is_lifetime else 'No',
            clean(pt) or 'NA', 'NA',
            f"PT: {pt or 'NA'} | Renewal: {renewal or 'NA'}",
        ]
        write_row(ws, HEADER_ROW + i, row, n_cols)

    # --------- 07A_Waiting_Period_Standard
    ws = init_sheet(wb, '07A_Waiting_Period_Standard', SCHEMA['07A_Waiting_Period_Standard'])
    n_cols = len(SCHEMA['07A_Waiting_Period_Standard'])
    for i, p in enumerate(products, 1):
        wp = feat(p['features'], 'Waiting Period')
        v, u = parse_waiting_period(wp)
        d = m = y = None
        if isinstance(v, int):
            if u == 'Day': d = v
            elif u == 'Month': m = v
            elif u == 'Year': y = v
        row = [
            p['_variant_id'], p['_variant_name'], p['product_name'], p['_company_short'],
            'Standard waiting period' if not is_na(wp) else 'NA',
            v if v is not None else 'NA', u or 'NA', d, m, y,
            'NA', 'NA', clean(wp) or 'NA',
        ]
        write_row(ws, HEADER_ROW + i, row, n_cols)

    # --------- 07B-07E (placeholder rows)
    for sn, label in [
        ('07B_Waiting_Period_Accident', 'Accident waiting period'),
        ('07C_Waiting_Period_Specific', 'Specific-condition waiting period'),
        ('07D_Waiting_Period_Special(90D)', 'Special 90-day waiting'),
        ('07E_Waiting_Period_Other(12MWP)', 'Other 12-month waiting'),
    ]:
        ws = init_sheet(wb, sn, SCHEMA[sn])
        n_cols = len(SCHEMA[sn])
        for i, p in enumerate(products, 1):
            wp = feat(p['features'], 'Waiting Period')
            row = [
                p['_variant_id'], p['_variant_name'], p['product_name'], p['_company_short'],
                'NA', 'NA', 'NA', None, None, None, 'NA', 'NA',
                f"Not separately specified in source. Overall WP: {wp or 'NA'}",
            ]
            write_row(ws, HEADER_ROW + i, row, n_cols)

    # --------- 08_Policy_Year_Benefit_Amt
    ws = init_sheet(wb, '08_Policy_Year_Benefit_Amt', SCHEMA['08_Policy_Year_Benefit_Amt'])
    n_cols = len(SCHEMA['08_Policy_Year_Benefit_Amt'])
    for i, p in enumerate(products, 1):
        max_si = feat(p['features'], 'Maximum Sum Insured (PHP)', 'Maximum Sum Insured',
                      'Max Sum Insured', 'Maximum Sum Assured')
        min_si = feat(p['features'], 'Minimum Sum Insured (PHP)', 'Minimum Sum Insured',
                      'Min Sum Insured', 'Minimum Sum Assured')
        val, ccy, _ = parse_amount(max_si)
        row = [
            p['_variant_id'], p['_variant_name'], p['product_name'], p['_company_short'],
            val if val is not None else 'NA', ccy or 'PHP', 'Sum Insured', 'NA',
            f"Max SI: {max_si or 'NA'} | Min SI: {min_si or 'NA'}",
        ]
        write_row(ws, HEADER_ROW + i, row, n_cols)

    # --------- 09_Hospital_Stay_Benefit_Amt
    ws = init_sheet(wb, '09_Hospital_Stay_Benefit_Amt', SCHEMA['09_Hospital_Stay_Benefit_Amt'])
    n_cols = len(SCHEMA['09_Hospital_Stay_Benefit_Amt'])
    for i, p in enumerate(products, 1):
        hi_amt = feat(p['features'], 'Hospital_Income_Amount', 'Hospital Income Amount')
        val, ccy, _ = parse_amount(hi_amt)
        row = [
            p['_variant_id'], p['_variant_name'], p['product_name'], p['_company_short'],
            val if val is not None else 'NA', ccy or ('PHP' if val is not None else 'NA'),
            'NA', 'NA',
            clean(hi_amt) or 'Not specified in CI benchmarking source',
        ]
        write_row(ws, HEADER_ROW + i, row, n_cols)

    # --------- 10_Area_Covered
    ws = init_sheet(wb, '10_Area_Covered', SCHEMA['10_Area_Covered'])
    n_cols = len(SCHEMA['10_Area_Covered'])
    for i, p in enumerate(products, 1):
        tm = feat(p['features'], 'Target Market')
        ccy = feat(p['features'], 'Policy Currency')
        if ccy and 'php' in str(ccy).lower():
            area = 'Philippines (Domestic)'
        elif ccy and ('thb' in str(ccy).lower() or 'baht' in str(ccy).lower()):
            area = 'Thailand (Domestic)'
        else:
            area = 'NA'
        row = [
            p['_variant_id'], p['_variant_name'], p['product_name'], p['_company_short'],
            area, 'NA', 'NA',
            f"Target Market: {tm or 'NA'} | Currency: {ccy or 'NA'}",
        ]
        write_row(ws, HEADER_ROW + i, row, n_cols)

    # --------- 11_Payment_Frequency
    ws = init_sheet(wb, '11_Payment_Frequency', SCHEMA['11_Payment_Frequency'])
    n_cols = len(SCHEMA['11_Payment_Frequency'])
    for i, p in enumerate(products, 1):
        ppt = feat(p['features'], 'Premium Payment Term', 'PPT')
        flex = feat(p['features'], 'Flexible Premium Payment Option- Description',
                    'Flexible Premium Payment')
        combined = f"{ppt or ''} {flex or ''}".lower()
        has_annual = 'annual' in combined or 'yearly' in combined
        has_semi = 'semi-annual' in combined or 'semi annual' in combined
        has_quarter = 'quarter' in combined
        has_monthly = 'monthly' in combined
        if not any([has_annual, has_semi, has_quarter, has_monthly]) and not is_na(ppt):
            has_annual = True
        modes = []
        if has_annual: modes.append('Annual')
        if has_semi: modes.append('Semi-Annual')
        if has_quarter: modes.append('Quarterly')
        if has_monthly: modes.append('Monthly')
        row = [
            p['_variant_id'], p['_variant_name'], p['product_name'], p['_company_short'],
            'Yes' if has_annual else 'No', 'Yes' if has_semi else 'No',
            'Yes' if has_quarter else 'No', 'Yes' if has_monthly else 'No',
            ', '.join(modes) if modes else 'NA', 'No', 'NA',
            f"PPT: {ppt or 'NA'}",
        ]
        write_row(ws, HEADER_ROW + i, row, n_cols)

    # --------- 12_Family_Discounts
    ws = init_sheet(wb, '12_Family_Discounts', SCHEMA['12_Family_Discounts'])
    n_cols = len(SCHEMA['12_Family_Discounts'])
    for i, p in enumerate(products, 1):
        fe_d = feat(p['features'], 'Family_Extension_Benefit_Description', 'Family Extension Benefit')
        fe_a = feat(p['features'], 'Family_Extension_Benefit_Amount', 'Family Extension Amount')
        has_fd = 'No' if is_na(fe_d) and is_na(fe_a) else 'Yes'
        write_row(ws, HEADER_ROW + i, [
            p['_variant_id'], p['_variant_name'], p['product_name'], p['_company_short'],
            has_fd, 'NA', 'NA', 'NA', 'NA',
            f"Family Extension: {fe_d or 'NA'} | Amount: {fe_a or 'NA'}",
        ], n_cols)

    # --------- 13-27 IPD/OPD placeholder sheets
    IPD_PLACEHOLDERS = [
        ('13_Room_Board_Inpatient', 'Room/Board IPD'),
        ('14_ICU_Room_Inpatient', 'ICU Room IPD'),
        ('15_Medical_Services_Fees', 'Medical Services / Take-home meds'),
        ('16_IPD_Physician_Fees', 'IPD Physician Fees'),
        ('17_Medical_Expenses_Surgery_Pro', 'Surgery procedure expenses'),
        ('18_Day_Surgery_Major', 'Day surgery major'),
        ('19_Diagnostic_Services_IPD', 'Diagnostic services IPD'),
        ('20_OPD_Accident_24_Hours', 'OPD Accident 24hr'),
        ('21_Rehabilitation_Expenses_IPD', 'Rehabilitation expenses IPD'),
        ('22_Dialysis_Treatment', 'Dialysis treatment'),
        ('23_Cancer_Treatment_Radiation', 'Cancer radiation treatment'),
        ('24_Chemo_Treatment_Expenses', 'Chemotherapy expenses'),
        ('25_Emergency_Ambulance_Services', 'Emergency ambulance'),
        ('26_Minor_Surgery_Expenses', 'Minor surgery expenses'),
        ('27_OPD_Medical_Services', 'OPD medical services'),
    ]
    for sn, label in IPD_PLACEHOLDERS:
        ws = init_sheet(wb, sn, SCHEMA[sn])
        n_cols = len(SCHEMA[sn])
        for i, p in enumerate(products, 1):
            row = [
                p['_variant_id'], p['_variant_name'], p['product_name'], p['_company_short'],
                label,
            ]
            while len(row) < n_cols - 1:
                row.append('NA')
            row.append('Not specified in Critical Illness benchmarking source')
            row = row[:n_cols]
            write_row(ws, HEADER_ROW + i, row, n_cols)

    # --------- 28_Critical_Illness_Benefits
    ws = init_sheet(wb, '28_Critical_Illness_Benefits', SCHEMA['28_Critical_Illness_Benefits'])
    n_cols = len(SCHEMA['28_Critical_Illness_Benefits'])
    conditions_list = cond_data['conditions'] if cond_data else []
    for i, p in enumerate(products, 1):
        desc = feat(p['features'], 'CI_Benefit_Description', 'CI Benefit Description')
        amt = feat(p['features'], 'CI_Benefit_Amount', 'CI Benefit Amount')
        cond = feat(p['features'], 'CI_Benefit_Conditions', 'CI Benefit Conditions')
        major_pct = feat(p['features'], 'Major CI Payout %', 'Major Payout %')
        minor_pct = feat(p['features'], 'Minor / Early-Stage CI Payout %', 'Minor Payout %')
        multi = feat(p['features'], 'Multiple Claims Allowed?', 'Multiple Claims')
        waiting = feat(p['features'], 'Waiting Period')
        term_after = feat(p['features'], 'Policy Terminates After CI Claim?', 'Terminates After CI')

        amt_s = str(amt or '').lower()
        pct = parse_pct(major_pct) or parse_pct(amt)
        is_pct = '%' in amt_s or 'face amount' in amt_s or 'sum assured' in amt_s
        amount_basis = 'NA'; amount_numeric = 'NA'; amount_multiplier = 'NA'; amount_ccy = 'NA'
        if is_pct:
            amount_basis = 'Percentage of SI'
            if pct is not None:
                amount_multiplier = f"{pct:.0%}"
                amount_numeric = pct
        else:
            v, c, _ = parse_amount(amt)
            if v is not None:
                amount_basis = 'Fixed Amount'
                amount_numeric = v
                amount_ccy = c or 'PHP'

        covered = 'Yes' if not is_na(desc) else 'No'
        missing = 'NA' if covered == 'Yes' else 'Not described in source'
        cov_count = 0
        if conditions_list:
            cov_count = sum(
                1 for x in conditions_list
                if str(x['coverage'].get(p['product_name'], '')).strip().lower() == 'covered'
            )
        scope = f"{cov_count} conditions covered (from condition benchmarking)" if cov_count > 0 else 'NA'
        max_cl = 'Single' if not is_na(term_after) and 'yes' in str(term_after).lower() \
                 else ('Multiple' if not is_na(multi) and 'yes' in str(multi).lower() else 'NA')
        trigger_stage = 'Major + Minor/Early' if (not is_na(minor_pct) or 'minor' in amt_s) else 'Major Only'

        row = [
            p['_variant_id'], p['_variant_name'], p['product_name'], p['_company_short'],
            'critical_illness_benefit',
            'Lump-sum CI benefit' if covered == 'Yes' else 'NA',
            clean(cond) or clean(desc) or 'NA',
            amount_multiplier if amount_multiplier != 'NA' else (amount_numeric if amount_numeric != 'NA' else 'Null'),
            amount_basis, 'Per policy term' if max_cl == 'Single' else 'Per claim',
            'Sum Assured' if is_pct else 'NA', clean(cond) or 'NA',
            f"Desc: {desc or 'NA'} | Amount: {amt or 'NA'} | Cond: {cond or 'NA'}",
            covered, missing, amount_basis, amount_numeric, amount_multiplier, amount_ccy,
            scope, scope, max_cl,
            'Per policy term' if max_cl == 'Single' else 'NA',
            trigger_stage, 'No',
            'Yes' if not is_na(waiting) else 'NA',
            'Yes' if not is_na(amt) and 'addition' in amt_s else 'No',
            'Yes' if covered == 'Yes' and not is_na(term_after) and 'yes' in str(term_after).lower() else 'No',
            'No' if max_cl == 'Single' else ('Yes' if max_cl == 'Multiple' else 'NA'),
            'Lump-sum' if covered == 'Yes' else 'NA',
            'Standard CI lump-sum' if covered == 'Yes' else 'NA',
            cov_count, f"{cov_count} covered conditions, {scope}",
            1 if covered == 'Yes' else 0,
            'CI core' if covered == 'Yes' else 'NA',
            amount_numeric if isinstance(amount_numeric, (int, float)) else 0,
        ]
        write_row(ws, HEADER_ROW + i, row, n_cols)

    # --------- 28A_CI_Conditions_Coverage
    if conditions_list:
        ws = init_sheet(wb, '28A_CI_Conditions_Coverage', SCHEMA['28A_CI_Conditions_Coverage'])
        n_cols = len(SCHEMA['28A_CI_Conditions_Coverage'])
        rix = HEADER_ROW
        for p in products:
            for cond in conditions_list:
                cov = cond['coverage'].get(p['product_name'])
                cov_s = str(cov).strip() if cov is not None else 'NA'
                if not cov_s or cov_s.lower() == 'none':
                    cov_s = 'NA'
                rix += 1
                write_row(ws, rix, [
                    p['_variant_id'], p['_variant_name'], p['product_name'], p['_company_short'],
                    cond['sno'], cond['category'], cond['name'],
                    cond['severity'], cond['stage'] or 'NA', cov_s,
                ], n_cols)

    # --------- 29_Misc_&_Extra_Benefits  +  per-benefit sheets (33+)
    EXTRA_BENEFITS = [
        ('Death Benefit', 'Death_Benefit_Description', 'Death_Benefit_Amount', 'Death_Benefit_Conditions'),
        ('Total Permanent Disability', 'TPD_Benefit_Description', 'TPD_Benefit_Amount', 'TPD_Benefit_Conditions'),
        ('Waiver of Premium', 'WaiverPremium_Description', 'WaiverPremium_Amount', 'WaiverPremium_Conditions'),
        ('Repatriation', 'Repatriation_Benefit_Description', 'Repatriation_Benefit_Amount', 'Repatriation_Benefit_Conditions'),
        ('Hospital Income', 'Hospital_Income_Description', 'Hospital_Income_Amount', 'Hospital_Income_Conditions'),
        ('Wellness Bonus', 'Wellness_Bonus_Description', 'Wellness_Bonus_Amount', 'Wellness_Bonus_Conditions'),
        ('Maturity Benefit', 'Maturity_Benefit_Description', 'Maturity_Benefit_Amount', 'Maturity_Benefit_Conditions'),
        ('Free Look Period', 'Free-Look_Period_Description', 'Free-Look_Period_Amount', 'Free-Look_Period_Conditions'),
        ('Gender Specific', 'Gender_Specific_Description', 'Gender_Specific_Amount', 'Gender_Specific_Conditions'),
        ('Recovery Benefit', 'Recovery_Benefit_Description', 'Recovery_Benefit_Amount', 'Recovery_Benefit_Conditions'),
        ('Accidental Death', 'Accidental_Death_Benefit_Description', 'Accidental_Death_Benefit_Amount', 'Accidental_Death_Benefit_Conditions'),
        ('Dividend', 'Dividend_Description', 'Dividend_Amount', 'Dividend_Conditions'),
        ('CI Buyback', 'CI_Buyback_Benefit', 'CI_Buyback_Amount', 'CI_Buyback_Conditions'),
        ('Cash Benefit', 'Cash_Benefit_Descriptioin', 'Cash_Benefit_Amount', 'Cash_Benefit_Conditions'),
        ('Treatment Support', 'Treatment_Support_Benefit_Description', 'Treatment_Support_Benefit_Amount', 'Treatment_Support_Benefit_Conditions'),
        ('Family Extension', 'Family_Extension_Benefit_Description', 'Family_Extension_Benefit_Amount', 'Family_Extension_Benefit_Conditions'),
        ('Accident & Disability', 'Accident & Disability Cover_Description', 'Accident & Disability Cover_Amount', 'Accident & Disability Cover_Conditions'),
        ('Digital Policy & Claims', 'Digital Policy & Claims_Description', 'Digital Policy & Claims_Amount', 'Digital Policy & Claims_Conditions'),
        ('Advanced CI Claim', 'Advanced_CI_Claim_Benefit_Description', 'Advanced_CI_Claim_Benefit_Amount', 'Advanced_CI_Claim_Benefit_Conditions'),
        ("Payor's Benefit", "Payor's_Benefit_Description", "Payor's_Benefit_Amount", "Payor's_Benefit_Conditions"),
        ('Fund Top-Up', 'Fund_Top-Up_Description', 'Fund_Top-Up_Amount', 'Fund_Top-Up_Conditions'),
        ('Fund Switching', 'Fund_Switching_Description', 'Fund_Switching_Amount', 'Fund_Switching_Conditions'),
        ('Loyalty Bonus', 'Loyalty_Bonus_Description', 'Loyalty_Bonus_Amount', 'Loyalty_Bonus_Conditions'),
        ('Emerging Condition Benefit', 'Emerging_Condition_Benefit_Description', 'Emerging_Condition_Benefit_Amount', 'Emerging_Condition_Benefit_Conditions'),
        ("Payor's Death Benefit", "Payour's_Death_Benefit_Description", "Payour's_Death_Benefit_Amount", "Payour's_Death_Benefit_Conditions"),
        ('Advanced Health Fund Withdrawal', 'Advanced Health Fund Withdrawal-Description', 'Advanced Health Fund Withdrawal - Benefit', 'Advanced Health Fund Withdrawal - Conditon'),
        ('Lifetime Coverage Extension', 'Lifetime Coverage Extension - Description', 'Lifetime Coverage Extension - Benefit', 'Lifetime Coverage Extension - Condition'),
        ('Flexible Premium Payment', 'Flexible Premium Payment Option- Description', 'Flexible Premium Payment Option- Benefit', 'Flexible Premium Payment Option- amount'),
    ]

    # --- summary sheet 29 (one row per product per benefit found) ---
    ws = init_sheet(wb, '29_Misc_&_Extra_Benefits', SCHEMA['29_Misc_&_Extra_Benefits'])
    n_cols = len(SCHEMA['29_Misc_&_Extra_Benefits'])
    rix = HEADER_ROW
    for p in products:
        for bname, dk, ak, ck in EXTRA_BENEFITS:
            d = feat(p['features'], dk)
            a = feat(p['features'], ak)
            c = feat(p['features'], ck)
            if is_na(d) and is_na(a) and is_na(c):
                continue
            v, ccy, _ = parse_amount(a)
            rix += 1
            write_row(ws, rix, [
                p['_variant_id'], p['_variant_name'], p['product_name'], p['_company_short'],
                bname, clean(d) or 'NA',
                'Yes' if not is_na(d) else 'NA',
                'Lump-sum' if v is not None else 'NA',
                v if v is not None else 'NA', ccy or 'NA',
                'NA', 'NA', 'NA', clean(c) or 'NA',
                f"Desc: {d or 'NA'} | Amount: {a or 'NA'} | Cond: {c or 'NA'}",
            ], n_cols)

    # --- 30, 31, 32 placeholder sheets (keep these numbers stable) ---
    for sn, label in [
        ('30_Deductible', 'Deductible'),
        ('31_Copayment', 'Copayment'),
        ('32_No_Claim_Bonus', 'No Claim Bonus'),
    ]:
        ws = init_sheet(wb, sn, SCHEMA[sn])
        n_cols = len(SCHEMA[sn])
        for i, p in enumerate(products, 1):
            row = [
                p['_variant_id'], p['_variant_name'], p['product_name'], p['_company_short'],
                label,
            ]
            while len(row) < n_cols - 1:
                row.append('NA')
            row.append('Not specified in Critical Illness benchmarking source')
            row = row[:n_cols]
            write_row(ws, HEADER_ROW + i, row, n_cols)

    # --- per-benefit sheets (33+): one sheet per benefit type that has data
    # Standard schema: Variant ID | Variant Name | Product Name | Company Name |
    #                  description | amount_value | amount_currency | conditions | raw_text
    benefit_headers = ID_COLS + [
        'description', 'amount_value', 'amount_currency', 'conditions', 'raw_text',
    ]
    benefit_n_cols = len(benefit_headers)
    sheet_num = 33
    per_benefit_sheets_created = []
    for bname, dk, ak, ck in EXTRA_BENEFITS:
        # Only create the sheet if AT LEAST ONE product has data for this benefit
        any_data = False
        for p in products:
            d = feat(p['features'], dk)
            a = feat(p['features'], ak)
            c = feat(p['features'], ck)
            if not (is_na(d) and is_na(a) and is_na(c)):
                any_data = True
                break
        if not any_data:
            continue

        sheet_name = make_sheet_name(sheet_num, bname)
        ws = init_sheet(wb, sheet_name, benefit_headers)
        for i, p in enumerate(products, 1):
            d = feat(p['features'], dk)
            a = feat(p['features'], ak)
            c = feat(p['features'], ck)
            v, ccy, _ = parse_amount(a)
            raw = f"Desc: {d or 'NA'} | Amount: {a or 'NA'} | Cond: {c or 'NA'}"
            row = [
                p['_variant_id'], p['_variant_name'], p['product_name'], p['_company_short'],
                clean(d) or 'NA',
                v if v is not None else (clean(a) or 'NA'),
                ccy or 'NA',
                clean(c) or 'NA',
                raw,
            ]
            write_row(ws, HEADER_ROW + i, row, benefit_n_cols)
        per_benefit_sheets_created.append((sheet_name, bname))
        sheet_num += 1

    # --------- Populate 00_Index
    populate_index(ws_idx, products, conditions_list, registry, per_benefit_sheets_created)

    # Tab colours
    set_tab_colors(wb)

    # Active sheet
    wb.active = wb.sheetnames.index('00_Index')

    return wb


# =====================================================================
# INDEX SHEET & STYLING
# =====================================================================
def populate_index(ws, products, conditions_list, registry, per_benefit_sheets=None):
    """Populate 00_Index with Arial 10 throughout, bold grey headers, no borders."""
    per_benefit_sheets = per_benefit_sheets or []
    ws.sheet_view.showGridLines = False

    # ---- Title ----
    ws.merge_cells('A1:E1')
    c = ws['A1']
    c.value = 'CI Benchmarking — Codified Comparison'
    c.font = ARIAL_10_BOLD
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = NO_BORDER
    for col in 'BCDE':
        cell = ws[f'{col}1']
        cell.fill = HDR_FILL
        cell.font = ARIAL_10_BOLD
        cell.border = NO_BORDER
    ws.row_dimensions[1].height = 22

    # ---- Summary stats ----
    rows = [
        ('Companies:', len(registry['unique_companies'])),
        ('Products:', len(products)),
        ('Variants:', f"{len(products)} (V01 per product)"),
        ('CI Conditions:', len(conditions_list) if conditions_list else 'None detected'),
    ]
    for i, (k, v) in enumerate(rows, 3):
        kc = ws.cell(row=i, column=1, value=k)
        kc.font = ARIAL_10_BOLD
        kc.border = NO_BORDER
        vc = ws.cell(row=i, column=2, value=v)
        vc.font = ARIAL_10
        vc.border = NO_BORDER

    # ---- Sheet Inventory section title ----
    section_row = 8
    ws.merge_cells(start_row=section_row, start_column=1, end_row=section_row, end_column=5)
    sc = ws.cell(row=section_row, column=1, value='Sheet Inventory')
    sc.font = ARIAL_10_BOLD
    sc.fill = HDR_FILL
    sc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    sc.border = NO_BORDER
    for col in range(2, 6):
        cell = ws.cell(row=section_row, column=col)
        cell.fill = HDR_FILL
        cell.font = ARIAL_10_BOLD
        cell.border = NO_BORDER
    ws.row_dimensions[section_row].height = 22

    # ---- Inventory column headers ----
    inv_headers = ['Sheet', 'Description', 'Rows', 'Source Coverage', 'Notes']
    hdr_row = 9
    for c_idx, h in enumerate(inv_headers, 1):
        cell = ws.cell(row=hdr_row, column=c_idx, value=h)
        cell.font = ARIAL_10_BOLD
        cell.fill = HDR_FILL
        cell.alignment = HDR_ALIGN
        cell.border = NO_BORDER
    ws.row_dimensions[hdr_row].height = 22

    # ---- Inventory rows ----
    inventory = [
        ('01_Company_Master', 'List of companies and IDs', len(registry['unique_companies']), 'Full', ''),
        ('02_Product_Master', 'Product catalogue with IDs', len(products), 'Full', ''),
        ('03_Variant_Master', 'Variants (one V01 per product)', len(products), 'Full', ''),
        ('04_Entry_Age', 'Min/max entry ages, codified', len(products), 'Good', 'From source Entry Age fields'),
        ('05_Policy_Term', 'Policy duration codified', len(products), 'Good', 'Parsed from free text'),
        ('06_Coverage', 'Coverage max age & renewal', len(products), 'Good', ''),
        ('07A_Waiting_Period_Standard', 'Standard waiting period', len(products), 'Partial', 'Source has one WP field'),
        ('07B-07E Waiting Periods', 'Accident / Specific / 90D / 12M variants', len(products), 'NA', 'Not separated in source'),
        ('08_Policy_Year_Benefit_Amt', 'Max benefit per year (SI)', len(products), 'Partial', 'Some products show NA for SI'),
        ('09_Hospital_Stay_Benefit_Amt', 'Per-stay max benefit', len(products), 'NA', 'CI source not focused on hospital'),
        ('10_Area_Covered', 'Geographic coverage', len(products), 'Inferred', 'From currency'),
        ('11_Payment_Frequency', 'Premium payment modes', len(products), 'Good', 'Parsed from PPT'),
        ('12_Family_Discounts', 'Family discount flag', len(products), 'Partial', 'From Family_Extension fields'),
        ('13-27 IPD/OPD benefits', 'Room/ICU/Surgery/etc.', len(products), 'NA', 'Source is CI-focused'),
        ('28_Critical_Illness_Benefits', 'Core CI benefit codification', len(products), 'STRONG', 'Best-mapped sheet'),
    ]
    if conditions_list:
        inventory.append(('28A_CI_Conditions_Coverage',
                          'Per-condition coverage (long format)',
                          len(products) * len(conditions_list), 'Full',
                          f"{len(products)} variants × {len(conditions_list)} conditions"))
    inventory.extend([
        ('29_Misc_&_Extra_Benefits', 'Summary of all extra benefits (per row)', '~28×products', 'STRONG', 'See per-benefit sheets below'),
        ('30_Deductible', 'Deductibles', len(products), 'NA', 'Not in CI source'),
        ('31_Copayment', 'Copayments', len(products), 'NA', 'Not in CI source'),
        ('32_No_Claim_Bonus', 'No-claim bonuses', len(products), 'NA', 'Not in CI source'),
    ])
    # Per-benefit sheets
    for sheet_name, benefit_label in per_benefit_sheets:
        inventory.append((sheet_name, benefit_label, len(products), 'Good',
                          'One row per product'))

    for i, (s, desc, rcount, cov, n) in enumerate(inventory):
        r = hdr_row + 1 + i
        for c_idx, val in enumerate([s, desc, rcount, cov, n], 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.font = ARIAL_10_BOLD if c_idx == 1 else ARIAL_10
            cell.alignment = Alignment(vertical='top', wrap_text=True,
                                       horizontal='center' if c_idx == 4 else 'left')
            cell.border = NO_BORDER

    # ---- Column widths ----
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 52
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 42


def set_tab_colors(wb):
    """No-op kept for compatibility; per the user's brief, no special tab colors are applied."""
    return


# =====================================================================
# MAIN
# =====================================================================
def codify(input_path: Path, output_path: Path) -> dict:
    """Run the full pipeline. Returns a summary dict."""
    wb_in = load_workbook(input_path, data_only=True)
    products, cond_data = extract_products(wb_in)
    if not products:
        raise RuntimeError('No products found in the input file. Check the Feature Benchmarking sheet structure.')

    wb_out = build_workbook(products, cond_data)
    wb_out.save(output_path)

    return {
        'input': str(input_path),
        'output': str(output_path),
        'companies': len({p['company_name'] for p in products}),
        'products': len(products),
        'conditions': len(cond_data['conditions']) if cond_data else 0,
        'sheets': len(wb_out.sheetnames),
    }


def main():
    parser = argparse.ArgumentParser(
        description='Convert a CI Benchmarking workbook into a CODIFY COMPARISON workbook.'
    )
    parser.add_argument('input', help='Path to the source benchmarking .xlsx file')
    parser.add_argument('output', nargs='?', default=None,
                        help='Optional output path. Defaults to "<input>_Codified.xlsx".')
    args = parser.parse_args()

    in_path = Path(args.input)
    if not in_path.exists():
        print(f"ERROR: input file not found: {in_path}", file=sys.stderr)
        sys.exit(1)
    out_path = Path(args.output) if args.output else in_path.with_name(in_path.stem + '_Codified.xlsx')

    print(f"→ Reading:  {in_path}")
    print(f"→ Writing:  {out_path}")
    try:
        summary = codify(in_path, out_path)
    except Exception as e:
        print(f"\nERROR: {e}", file=sys.stderr)
        sys.exit(2)

    print()
    print(f"✓ Done.")
    print(f"  Companies: {summary['companies']}")
    print(f"  Products:  {summary['products']}")
    print(f"  CI conditions: {summary['conditions']}")
    print(f"  Output sheets: {summary['sheets']}")


if __name__ == '__main__':
    main()
